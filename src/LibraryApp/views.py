from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.db.models import Count, Q
from datetime import datetime, timedelta
from .models import BankAccount, Reader, ReaderType, Parameter, BookTitle, Author, BookImportReceipt, BookImportDetail, Book, AuthorDetail, BookItem, BorrowReturnReceipt, Receipt, Category, UserGroup, Function, Permission
from .forms import ReaderForm, LibraryLoginForm, BookImportForm, BookImportExcelForm, BookSearchForm, BorrowBookForm, ReturnBookForm, ReceiptForm, ParameterForm, BookEditForm, ReaderTypeForm, UserGroupForm, FunctionForm
from .decorators import manager_required, staff_required, permission_required


def home_view(request):
    """Trang chủ hệ thống với dashboard statistics"""
    context = {}
    
    if request.user.is_authenticated:
        # Đếm tổng số độc giả
        context['total_readers'] = Reader.objects.filter(is_active=True).count()
        
        # Đếm tổng số đầu sách (BookTitle)
        context['total_books'] = BookTitle.objects.count()
        
        # Đếm số phiếu mượn đang hoạt động (chưa trả, loại trừ đã hủy)
        context['active_borrows'] = BorrowReturnReceipt.objects.filter(
            return_date__isnull=True,
            is_cancelled=False  # Loại trừ phiếu đã hủy
        ).count()
        
        # Đếm số sách quá hạn (loại trừ đã hủy)
        from datetime import date
        context['overdue_books'] = BorrowReturnReceipt.objects.filter(
            return_date__isnull=True,
            is_cancelled=False,  # Loại trừ phiếu đã hủy
            due_date__lt=date.today()
        ).count()
    
    return render(request, 'app/home.html', context)


# ==================== AUTHENTICATION ====================

def login_view(request):
    """
    Đăng nhập hệ thống
    Dành cho: SuperUser, Staff (Quản lý, Thủ thư)
    """
    # Nếu đã đăng nhập, chuyển về trang chủ
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        form = LibraryLoginForm(request, data=request.POST)
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            # Case-insensitive lookup
            user_obj = User.objects.filter(username__iexact=username).first()
            if not user_obj:
                # User not found. Let form validation handle the error generation naturally.
                pass
            else:
                # Check inactive status first - BLOCK before any further processing
                if not user_obj.is_active:
                    # Clear any existing errors and set only our message
                    from django.forms.utils import ErrorList
                    _ = form.errors  # Trigger validation
                    form.errors['__all__'] = ErrorList(['Tài khoản chưa được kích hoạt. Vui lòng liên hệ Admin.'])
                else:
                    try:
                        # Logic for LibraryUser (tracking attempts)
                        lib_user = user_obj.library_user
                        
                        # Check lockout
                        if lib_user.failed_login_attempts >= 5:
                            from django.forms.utils import ErrorList
                            _ = form.errors
                            form.errors['__all__'] = ErrorList(['Tài khoản đã bị khoá do đăng nhập sai quá 5 lần. Vui lòng liên hệ Admin hoặc sử dụng chức năng "Quên mật khẩu" để lấy lại mật khẩu.'])
                        else:
                            # Attempt authentication with canonical username
                            user = authenticate(request, username=user_obj.username, password=password)
                            
                            if user:
                                # Success -> Reset counter
                                lib_user.failed_login_attempts = 0
                                lib_user.save(update_fields=['failed_login_attempts'])
                                login(request, user)
                                
                                # Welcome message
                                if user.is_superuser:
                                    messages.success(request, f'Chào mừng Quản trị viên {user.username}! Bạn có toàn quyền truy cập.')
                                elif user.is_staff:
                                    messages.success(request, f'Chào mừng {user.username}!')
                                else:
                                    messages.warning(request, 'Tài khoản này không có quyền truy cập hệ thống.')
                                    logout(request)
                                    return redirect('login')
                                
                                next_url = request.GET.get('next', 'home')
                                return redirect(next_url)
                            else:
                                # Failed password -> Increment counter
                                lib_user.failed_login_attempts += 1
                                lib_user.save(update_fields=['failed_login_attempts'])
                                
                                remaining = 5 - lib_user.failed_login_attempts
                                from django.forms.utils import ErrorList
                                _ = form.errors  # Trigger validation
                                
                                if remaining <= 0:
                                    form.errors['__all__'] = ErrorList(['Tài khoản đã bị khoá do đăng nhập sai quá 5 lần. Vui lòng liên hệ Admin hoặc sử dụng chức năng "Quên mật khẩu" để lấy lại mật khẩu.'])
                                else:
                                    # Custom message: just "Wrong password..."
                                    form.errors['__all__'] = ErrorList([f"Sai mật khẩu. Bạn còn {remaining} lần thử."])
                    
                    except (AttributeError, ObjectDoesNotExist):
                        # Fallback for users without LibraryUser (e.g. pure superuser)
                        user = authenticate(request, username=user_obj.username, password=password)
                        if user:
                            login(request, user)
                            if user.is_superuser:
                                messages.success(request, f'Chào mừng Quản trị viên {user.username}! Bạn có toàn quyền truy cập.')
                            elif user.is_staff:
                                messages.success(request, f'Chào mừng {user.username}!')
                            else:
                                messages.warning(request, 'Tài khoản này không có quyền truy cập hệ thống.')
                                logout(request)
                                return redirect('login')
                            
                            next_url = request.GET.get('next', 'home')
                            return redirect(next_url)
    else:
        form = LibraryLoginForm()
    
    context = {
        'form': form,
        'page_title': 'Đăng nhập'
    }
    
    return render(request, 'accounts/login.html', context)


def logout_view(request):
    """Đăng xuất khỏi hệ thống"""
    if request.user.is_authenticated:
        username = request.user.username
        logout(request)
        messages.success(request, f'Đã đăng xuất tài khoản {username}.')
    return redirect('/')


from django.contrib.auth.views import PasswordResetView
from django.urls import reverse_lazy
from django.conf import settings
from urllib.parse import urlparse
from .forms import UserEmailPasswordResetForm

class CustomPasswordResetView(PasswordResetView):
    template_name = 'accounts/password_reset.html'
    email_template_name = 'accounts/password_reset_email.html'
    subject_template_name = 'accounts/password_reset_subject.txt'
    success_url = reverse_lazy('password_reset_done')
    form_class = UserEmailPasswordResetForm
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Quên mật khẩu'
        return context

    def form_valid(self, form):
        # Determine which URL info to use
        # Logic: 
        # 1. Start with LOCAL_URL as default
        # 2. If PUBLIC_URL is defined AND (we are not in DEBUG OR request host matches public domain), use PUBLIC_URL
        
        target_url = settings.LOCAL_URL
        public_url = settings.PUBLIC_URL
        
        if public_url:
            # Parse public url to get domain
            try:
                public_parts = urlparse(public_url)
                request_host = self.request.get_host()
                
                # If the current request is coming from the public domain, OR we strictly want to valid public url usage
                # A simple heuristic: if request host is part of public url, use it.
                # However, user said: "If tunnel...". Tunnel often means request host IS the public domain.
                # But sometimes proxy setup is tricky. 
                # Let's trust the user's intent: If PUBLIC_URL is set, we prefer it if we are accessing via that domain
                # OR if DEBUG is False (Production/Tunnel mode assumption).
                
                if public_parts.netloc in request_host or not settings.DEBUG:
                    target_url = public_url
            except:
                pass
        
        # Extract domain and protocol from target_url
        parsed = urlparse(target_url)
        domain = parsed.netloc
        scheme = parsed.scheme
        use_https = (scheme == 'https')
        
        opts = {
            'use_https': use_https,
            'token_generator': self.token_generator,
            'from_email': self.from_email,
            'email_template_name': self.email_template_name,
            'subject_template_name': self.subject_template_name,
            'request': self.request,
            'html_email_template_name': self.html_email_template_name,
            'extra_email_context': self.extra_email_context,
            'domain_override': domain,
        }
        form.save(**opts)
        return redirect(self.get_success_url())

# Rename to match urls.py expectation or update urls.py
password_reset_view = CustomPasswordResetView.as_view()


@login_required
def user_profile_view(request):
    """
    Trang cá nhân của người dùng
    Cho phép xem và cập nhật thông tin cá nhân
    """
    user = request.user
    
    if request.method == 'POST':
        # Cập nhật thông tin cơ bản
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        
        # New fields
        phone_number = request.POST.get('phone_number', '').strip()
        address = request.POST.get('address', '').strip()
        dob_str = request.POST.get('date_of_birth')
        position = request.POST.get('position', '').strip()
        
        # Validate email
        if email and User.objects.filter(email=email).exclude(id=user.id).exists():
            messages.error(request, 'Email đã được sử dụng bởi người dùng khác.')
        else:
            user.first_name = first_name
            user.last_name = last_name
            if email:
                user.email = email
            user.save()
            
            # Update LibraryUser
            from .models import LibraryUser
            from django.utils import timezone
            
            try:
                library_user = user.library_user
                
                library_user.full_name = f"{first_name} {last_name}".strip() or user.username
                library_user.phone_number = phone_number
                library_user.address = address
                
                if dob_str:
                    try:
                        library_user.date_of_birth = timezone.datetime.strptime(dob_str, '%Y-%m-%d').date()
                    except ValueError:
                        pass
                
                if user.is_superuser and position: # Only admin can edit position
                    library_user.position = position
                    
                library_user.save()
            except LibraryUser.DoesNotExist:
                pass # Should not happen for library users
                
            messages.success(request, 'Cập nhật thông tin thành công!')
        
        # Đổi mật khẩu nếu có
        old_password = request.POST.get('old_password', '')
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')
        
        if old_password and new_password:
            if not user.check_password(old_password):
                messages.error(request, 'Mật khẩu cũ không đúng.')
            elif new_password != confirm_password:
                messages.error(request, 'Mật khẩu mới và xác nhận không khớp.')
            elif len(new_password) < 6:
                messages.error(request, 'Mật khẩu mới phải có ít nhất 6 ký tự.')
            else:
                user.set_password(new_password)
                user.save()
                # Re-login để giữ session
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, user)
                messages.success(request, 'Đổi mật khẩu thành công!')
        
        return redirect('user_profile')
    
    context = {
        'page_title': 'Trang cá nhân',
        'profile_user': user,
    }
    
    return render(request, 'accounts/profile.html', context)


# ==================== YC1: LẬP THẺ ĐỘC GIẢ ====================

@permission_required('Lập thẻ độc giả', 'add')
def reader_create_view(request):
    """
    View lập thẻ độc giả mới - YC1
    
    Business Rules (QĐ1):
    - Tuổi từ min_age đến max_age (mặc định 18-55)
    - Phải chọn loại độc giả hợp lệ
    - Thẻ có giá trị theo card_validity_period (mặc định 6 tháng)
    """
    
    # Kiểm tra hệ thống đã được cấu hình chưa
    params = Parameter.objects.first()
    if not params:
        messages.error(request, 'Hệ thống chưa được cấu hình. Vui lòng liên hệ quản trị viên.')
        return redirect('home')
    
    if request.method == 'POST':
        form = ReaderForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Lưu độc giả (auto-calculate expiration_date trong model.save())
                    reader = form.save()
                    
                    messages.success(
                        request,
                        f'Lập thẻ độc giả thành công! '
                        f'Mã độc giả: {reader.id} - {reader.reader_name}. '
                        f'Thẻ có hiệu lực đến: {reader.expiration_date.strftime("%d/%m/%Y")}'
                    )
                    return redirect('reader_detail', reader_id=reader.id)
            except Exception as e:
                messages.error(request, f'Có lỗi xảy ra: {str(e)}')
        else:
            messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = ReaderForm()
    
    context = {
        'form': form,
        'params': params,
        'reader_types': ReaderType.objects.all(),
        'page_title': 'Lập thẻ độc giả mới'
    }
    
    return render(request, 'app/readers/reader_create.html', context)


@permission_required('Quản lý độc giả', 'change')
def reader_edit_view(request, reader_id):
    """
    View chỉnh sửa thông tin độc giả
    
    Cho phép sửa:
    - Họ tên, email, địa chỉ
    - Loại độc giả
    - Ngày sinh (với validation tuổi)
    - Ngày lập thẻ (sẽ tự động tính lại ngày hết hạn)
    """
    reader = get_object_or_404(Reader, id=reader_id)
    params = Parameter.objects.first()
    
    if not params:
        messages.error(request, 'Hệ thống chưa được cấu hình. Vui lòng liên hệ quản trị viên.')
        return redirect('home')
    
    if request.method == 'POST':
        form = ReaderForm(request.POST, instance=reader)
        if form.is_valid():
            try:
                with transaction.atomic():
                    reader = form.save()
                    
                    messages.success(
                        request,
                        f'Cập nhật thông tin độc giả thành công! '
                        f'Mã độc giả: {reader.id} - {reader.reader_name}.'
                    )
                    return redirect('reader_detail', reader_id=reader.id)
            except Exception as e:
                messages.error(request, f'Có lỗi xảy ra: {str(e)}')
        else:
            messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = ReaderForm(instance=reader)
    
    context = {
        'form': form,
        'params': params,
        'reader_types': ReaderType.objects.all(),
        'reader': reader,
        'is_edit': True,
        'page_title': f'Chỉnh sửa độc giả - {reader.reader_name}'
    }
    
    return render(request, 'app/readers/reader_create.html', context)

@permission_required('Quản lý độc giả', 'view')
def reader_detail_view(request, reader_id):
    """
    Xem chi tiết thẻ độc giả - Hiển thị thông tin sau khi lập thẻ
    """
    reader = get_object_or_404(Reader, id=reader_id)
    
    # Check permissions for actions
    from .decorators import check_permission
    can_edit_readers = check_permission(request.user, 'Quản lý độc giả', 'change')
    can_add_readers = check_permission(request.user, 'Lập thẻ độc giả', 'add')
    can_delete_readers = check_permission(request.user, 'Quản lý độc giả', 'delete')
    
    context = {
        'reader': reader,
        'can_edit_readers': can_edit_readers,
        'can_add_readers': can_add_readers,
        'can_delete_readers': can_delete_readers,
        'page_title': f'Thẻ độc giả - {reader.reader_name}'
    }
    
    return render(request, 'app/readers/reader_detail.html', context)


@permission_required('Quản lý độc giả', 'view')
def reader_list_view(request):
    """
    Danh sách độc giả - Để tra cứu và quản lý
    """
    readers = Reader.objects.select_related('reader_type').all().order_by('-card_creation_date')
    
    # Filter theo loại độc giả nếu có
    reader_type_id = request.GET.get('reader_type')
    if reader_type_id:
        readers = readers.filter(reader_type_id=reader_type_id)
    
    # Search theo tên hoặc email
    search_query = request.GET.get('search')
    if search_query:
        readers = readers.filter(
            reader_name__icontains=search_query
        ) | readers.filter(
            email__icontains=search_query
        )
    
    context = {
        'readers': readers,
        'reader_types': ReaderType.objects.all(),
        'page_title': 'Danh sách độc giả'
    }
    
    return render(request, 'app/readers/reader_list.html', context)


@permission_required('Quản lý độc giả', 'delete')
def reader_delete_view(request, reader_id):
    """
    View xóa độc giả
    
    Business Rules:
    - Chỉ xóa được nếu độc giả không còn sách đang mượn
    - Chỉ xóa được nếu độc giả không có nợ tiền phạt
    - Yêu cầu quyền DELETE trong chức năng "Quản lý độc giả"
    """
    reader = get_object_or_404(Reader, id=reader_id)
    
    if request.method == 'POST':
        try:
            # Kiểm tra sách đang mượn
            borrowing_count = BorrowReturnReceipt.objects.filter(
                reader=reader,
                return_date__isnull=True
            ).count()
            
            if borrowing_count > 0:
                messages.error(
                    request,
                    f'Không thể xóa độc giả "{reader.reader_name}" vì còn {borrowing_count} sách đang mượn. '
                    'Vui lòng trả hết sách trước khi xóa.'
                )
                return redirect('reader_detail', reader_id=reader.id)
            
            # Kiểm tra nợ tiền phạt
            if reader.total_debt > 0:
                messages.error(
                    request,
                    f'Không thể xóa độc giả "{reader.reader_name}" vì còn nợ {reader.total_debt:,}đ tiền phạt. '
                    'Vui lòng thu hết nợ trước khi xóa.'
                )
                return redirect('reader_detail', reader_id=reader.id)
            
            # Thực hiện xóa
            reader_name = reader.reader_name
            reader.delete()
            
            messages.success(
                request,
                f'Đã xóa độc giả "{reader_name}" thành công!'
            )
            return redirect('reader_list')
            
        except Exception as e:
            messages.error(request, f'Có lỗi xảy ra khi xóa độc giả: {str(e)}')
            return redirect('reader_detail', reader_id=reader.id)
    
    # GET: Hiển thị trang xác nhận xóa
    # Đếm số phiếu mượn và phiếu thu liên quan
    # Thống kê (loại trừ phiếu đã hủy)
    total_readers = Reader.objects.count()
    total_books = BookTitle.objects.count()
    total_borrowed = BorrowReturnReceipt.objects.filter(
        return_date__isnull=True,
        is_cancelled=False  # Loại trừ phiếu đã hủy
    ).count()
    total_overdue = BorrowReturnReceipt.objects.filter(
        return_date__isnull=True,
        is_cancelled=False,  # Loại trừ phiếu đã hủy
        due_date__lt=timezone.now()
    ).count()
    
    borrow_count = BorrowReturnReceipt.objects.filter(reader=reader).count()
    receipt_count = Receipt.objects.filter(reader=reader).count()
    borrowing_count = BorrowReturnReceipt.objects.filter(
        reader=reader,
        return_date__isnull=True
    ).count()
    
    context = {
        'reader': reader,
        'borrow_count': borrow_count,
        'receipt_count': receipt_count,
        'borrowing_count': borrowing_count,
        'can_delete': borrowing_count == 0 and reader.total_debt == 0,
        'page_title': f'Xóa độc giả - {reader.reader_name}'
    }
    
    return render(request, 'app/readers/reader_delete_confirm.html', context)



# ==================== YC2: TIẾP NHẬN SÁCH MỚI ====================

@permission_required('Lập phiếu nhập sách', 'add')
def book_import_select_view(request):
    """
    Trang trung gian - Chọn phương thức nhập sách
    Cho phép user chọn: Nhập thủ công hoặc Nhập từ Excel
    """
    context = {
        'page_title': 'Tiếp nhận sách mới'
    }
    return render(request, 'app/books/book_import_select.html', context)


@permission_required('Lập phiếu nhập sách', 'add')
def book_import_view(request):
    """
    View tiếp nhận sách mới - YC2
    
    Business Rules (QĐ2):
    - Chỉ nhận sách xuất bản trong vòng 8 năm (book_return_period)
    - Thể loại phải hợp lệ
    - Tác giả phải hợp lệ
    - Tạo phiếu nhập và chi tiết phiếu nhập
    """
    
    # Kiểm tra hệ thống đã được cấu hình chưa
    params = Parameter.objects.first()
    if not params:
        messages.error(request, 'Hệ thống chưa được cấu hình. Vui lòng liên hệ quản trị viên.')
        return redirect('home')
    
    if request.method == 'POST':
        form = BookImportForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Lấy dữ liệu từ form
                    book_title_name = form.cleaned_data['book_title']
                    category = form.cleaned_data['category']
                    
                    # Đọc tác giả từ HIDDEN INPUTS (primary source - reliable)
                    # Django form field cho SelectMultiple có thể không hoạt động đúng khi element bị ẩn
                    authors = []
                    new_authors_str = ''
                    
                    # 1. Đọc existing author IDs từ hidden input (primary)
                    author_ids_str = request.POST.get('author_ids', '').strip()
                    print(f"DEBUG Python: author_ids from hidden input: '{author_ids_str}'")
                    
                    if author_ids_str:
                        author_ids = [aid.strip() for aid in author_ids_str.split(',') if aid.strip() and aid.strip().isdigit()]
                        print(f"DEBUG Python: Parsed author IDs: {author_ids}")
                        for aid in author_ids:
                            try:
                                author = Author.objects.get(id=int(aid))
                                if author not in authors:
                                    authors.append(author)
                                    print(f"DEBUG Python: Added existing author: {author.author_name}")
                            except Author.DoesNotExist:
                                print(f"DEBUG Python: Author ID {aid} not found in database")
                    
                    # 2. Nếu hidden input rỗng, fallback về Django form field
                    if not authors:
                        form_authors = list(form.cleaned_data.get('authors', []))
                        print(f"DEBUG Python: Fallback to form.cleaned_data authors: {form_authors}")
                        authors = form_authors
                    
                    # 3. Đọc new authors từ hidden input (primary) hoặc form field
                    new_authors_str = request.POST.get('new_author_names', '').strip()
                    print(f"DEBUG Python: new_author_names from hidden input: '{new_authors_str}'")
                    
                    if not new_authors_str:
                        new_authors_str = form.cleaned_data.get('new_authors', '').strip()
                        print(f"DEBUG Python: Fallback to form.cleaned_data new_authors: '{new_authors_str}'")
                    # VALIDATION: Phải có ít nhất 1 tác giả
                    if not authors and not new_authors_str:
                        messages.error(request, 'Tác giả: Vui lòng chọn hoặc nhập ít nhất 1 tác giả')
                        from datetime import date
                        context = {
                            'form': form,
                            'params': params,
                            'min_year': date.today().year - params.book_return_period,
                            'page_title': 'Tiếp nhận sách mới'
                        }
                        return render(request, 'app/books/book_import.html', context)
                    
                    publish_year = form.cleaned_data['publish_year']
                    publisher = form.cleaned_data['publisher']
                    isbn = form.cleaned_data.get('isbn', '')
                    edition = form.cleaned_data.get('edition', '')
                    language = form.cleaned_data['language']
                    quantity = form.cleaned_data['quantity']
                    unit_price = form.cleaned_data['unit_price']
                    description = form.cleaned_data.get('description', '')
                    notes = form.cleaned_data.get('notes', '')
                    import_date = form.cleaned_data['import_date']
                    
                    # Xử lý tác giả mới
                    # JavaScript sử dụng delimiter '|||' để tránh conflict với tên có dấu phẩy
                    if new_authors_str:
                        # Support both delimiters: '|||' (new) and ',' (fallback for old submissions)
                        if '|||' in new_authors_str:
                            new_author_names = [name.strip() for name in new_authors_str.split('|||') if name.strip()]
                        else:
                            new_author_names = [name.strip() for name in new_authors_str.split(',') if name.strip()]
                        
                        for name in new_author_names:
                            # Safety: Strip 'new_' prefix if it somehow got through
                            if name.startswith('new_'):
                                name = name[4:].strip()
                            
                            # Validate: skip if name is empty after stripping
                            if not name:
                                continue
                            
                            # Validate: skip if name looks like an author ID (numeric only)
                            if name.isdigit():
                                # This is likely an ID that was incorrectly synced, skip it
                                print(f"DEBUG: Skipping numeric-only author name '{name}' - likely an ID")
                                continue
                            
                            # Tìm hoặc tạo tác giả mới (case-insensitive search để tránh trùng lặp)
                            existing_author = Author.objects.filter(author_name__iexact=name).first()
                            if existing_author:
                                author = existing_author
                            else:
                                author = Author.objects.create(author_name=name)
                            
                            if author not in authors:
                                authors.append(author)
                    
                    # Kiểm tra hoặc tạo BookTitle (case-insensitive)
                    existing_book_title = BookTitle.objects.filter(
                        book_title__iexact=book_title_name,
                        category=category
                    ).first()
                    if existing_book_title:
                        book_title = existing_book_title
                        # Cập nhật description nếu chưa có và có description mới
                        if description and not book_title.description:
                            book_title.description = description
                            book_title.save()
                    else:
                        book_title = BookTitle.objects.create(
                            book_title=book_title_name,
                            category=category,
                            description=description
                        )
                    
                    # Thêm tác giả cho BookTitle (cả cũ và mới)
                    # Nếu BookTitle đã tồn tại, ta vẫn nên kiểm tra xem có cần bổ sung tác giả không?
                    # Logic hiện tại: chỉ thêm nếu created=True. 
                    # Nếu BookTitle đã có, ta có thể muốn update thêm tác giả nếu chưa có?
                    # Để an toàn và nhất quán với yêu cầu, ta sẽ add thêm tác giả vào BookTitle nếu chưa có logic đó.
                    # Nhưng code cũ chỉ thêm khi created. Ta sẽ giữ nguyên logic đó hoặc cải tiến.
                    # Quyết định: Cập nhật connection tác giả cho BookTitle dù mới hay cũ
                    
                    current_authors = book_title.authors.all()
                    for author in authors:
                        if author not in current_authors:
                            AuthorDetail.objects.create(
                                author=author,
                                book_title=book_title
                            )
                    
                    # Smart duplicate detection (case-insensitive publisher):
                    # Find existing books with same title, publisher, year
                    existing_books = Book.objects.filter(
                        book_title=book_title,
                        publish_year=publish_year,
                        publisher__iexact=publisher
                    )
                    
                    # Check ISBN and Edition - both must match for same book
                    book = None
                    if isbn:
                        existing_books = existing_books.filter(isbn=isbn)
                    
                    # Different edition = different book (e.g., "Tái bản lần 1" vs "Tái bản lần 2")
                    if edition:
                        existing_books = existing_books.filter(edition=edition)
                    else:
                        # Only match books with no edition specified
                        existing_books = existing_books.filter(edition__isnull=True) | existing_books.filter(edition='')
                    
                    # If we found exact matches, use the first one (update quantity)
                    if existing_books.exists():
                        book = existing_books.first()
                        # Book exists with same details → will update quantity via BookImportDetail
                    else:
                        # No exact match → create new Book record
                        # Set quantity=1 initially (min validator requirement)
                        # BookImportDetail.save() will handle the actual quantity update
                        book = Book.objects.create(
                            book_title=book_title,
                            publish_year=publish_year,
                            publisher=publisher,
                            isbn=isbn if isbn else None,
                            quantity=1,
                            remaining_quantity=1,
                            unit_price=unit_price,
                            edition=edition,
                            language=language
                        )
                        # Mark as just created so we know to set quantity properly
                        book._just_created = True
                    # Note: BookImportDetail.save() handles quantity increment
                    # and BookItem creation automatically
                    
                    # Tạo phiếu nhập
                    # Kết hợp ngày từ form với giờ hiện tại để có datetime đầy đủ
                    from datetime import datetime, time
                    import_date_with_time = timezone.make_aware(
                        datetime.combine(import_date, timezone.localtime().time())
                    ) if import_date else timezone.now()
                    
                    receipt = BookImportReceipt.objects.create(
                        import_date=import_date_with_time,
                        created_by=request.user.username,
                        notes=notes
                    )
                    
                    # Tạo chi tiết phiếu nhập
                    import_detail = BookImportDetail.objects.create(
                        receipt=receipt,
                        book=book,
                        quantity=quantity,
                        unit_price=unit_price
                    )
                    
                    # Thông báo thành công
                    messages.success(
                        request,
                        f'Tiếp nhận sách thành công! '
                        f'Tựa sách: {book_title.book_title} '
                        f'({quantity} cuốn, {unit_price:,}đ/cuốn). '
                        f'Phiếu nhập #{receipt.id}.'
                    )
                    return redirect('book_import_detail', import_id=receipt.id)
            except ValidationError as e:
                # Lỗi validation từ model
                messages.error(request, f'Lỗi validation: {e.message if hasattr(e, "message") else str(e)}')
            except Exception as e:
                # Lỗi hệ thống khác
                messages.error(request, f'Có lỗi xảy ra: {str(e)}')
        else:
            # Hiển thị lỗi validation cho user
            for field, errors in form.errors.items():
                # Lấy label thân thiện từ form field
                field_label = form.fields[field].label if field in form.fields else field
                for error in errors:
                    messages.error(request, f'{field_label}: {error}')
    else:
        form = BookImportForm()
    
    # Lấy tham số để hiển thị thông tin QĐ2
    from datetime import date
    min_year = date.today().year - params.book_return_period
    
    # Lấy danh sách nhà xuất bản có sẵn từ DB
    existing_publishers = Book.objects.values_list('publisher', flat=True).distinct().order_by('publisher')
    existing_publishers = [p for p in existing_publishers if p]  # Filter out empty values
    
    context = {
        'form': form,
        'params': params,
        'min_year': min_year,
        'existing_publishers': existing_publishers,
        'page_title': 'Tiếp nhận sách mới'
    }
    
    return render(request, 'app/books/book_import.html', context)


@permission_required('Lập phiếu nhập sách', 'add')
def book_import_excel_view(request):
    """
    View nhập sách từ Excel
    """
    if request.method == 'POST':
        form = BookImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                import pandas as pd
                df = pd.read_excel(excel_file)
                
                # Normalize columns to lower case/strip
                df.columns = df.columns.astype(str).str.lower().str.strip()
                
                # Expected columns mapping
                # Tên sách | Thể loại | Tác giả | Năm XB | NXB | Số lượng | Đơn giá | Ngày nhập | ISBN | Phiên bản | Ngôn ngữ | Mô tả
                
                success_count = 0
                error_count = 0
                errors = []
                
                with transaction.atomic():
                    # Create one receipt for the whole batch? Or one per row?
                    # Let's create one receipt for the whole import file
                    from django.utils import timezone
                    
                    # Assume import date is today if not specified or mixed. 
                    # If file has 'Ngày nhập', use it. But receipt has one date.
                    # Strategy: One receipt per import session (now).
                    receipt = BookImportReceipt.objects.create(
                        import_date=timezone.now(),
                        created_by=request.user.username,
                        notes=f"Nhập khẩu lô hàng từ file {excel_file.name}"
                    )
                    
                    for index, row in df.iterrows():
                        try:
                            # 1. Parse Data
                            book_title_name = str(row.get('tên sách', '')).strip()
                            if not book_title_name or pd.isna(row.get('tên sách')):
                                continue # Skip empty rows
                                
                            category_name = str(row.get('thể loại', '')).strip()
                            author_names_str = str(row.get('tác giả', '')).strip()
                            
                            publish_year = row.get('năm xb', 2025)
                            publisher = str(row.get('nxb', '')).strip()
                            
                            quantity = row.get('số lượng', 1)
                            unit_price = row.get('đơn giá', 0)
                            
                            # Optional fields
                            isbn = str(row.get('isbn', '')).strip()
                            if isbn == 'nan': isbn = ''
                            
                            edition = str(row.get('phiên bản', '')).strip()
                            if edition == 'nan': edition = ''
                                
                            language = str(row.get('ngôn ngữ', 'Tiếng Việt')).strip()
                            if language == 'nan': language = 'Tiếng Việt'
                            
                            description = str(row.get('mô tả', '')).strip()
                            if description == 'nan': description = ''
                            
                            # Clean numbers
                            try: publish_year = int(publish_year)
                            except: publish_year = 2025
                            
                            try: quantity = int(quantity)
                            except: quantity = 1
                            if quantity < 1: quantity = 1
                            
                            try: unit_price = int(unit_price)
                            except: unit_price = 0
                            
                            # 2. Get/Create Category (case-insensitive)
                            existing_category = Category.objects.filter(category_name__iexact=category_name).first()
                            if existing_category:
                                category = existing_category
                            else:
                                category = Category.objects.create(category_name=category_name)
                            
                            # 3. Get/Create BookTitle (case-insensitive)
                            existing_book_title = BookTitle.objects.filter(
                                book_title__iexact=book_title_name,
                                category=category
                            ).first()
                            if existing_book_title:
                                book_title = existing_book_title
                                # Cập nhật description nếu chưa có và có description mới từ Excel
                                if description and not book_title.description:
                                    book_title.description = description
                                    book_title.save()
                            else:
                                book_title = BookTitle.objects.create(
                                    book_title=book_title_name,
                                    category=category,
                                    description=description
                                )
                            
                            # 4. Handle Authors (case-insensitive)
                            author_names = [name.strip() for name in author_names_str.split(',') if name.strip()]
                            current_authors = book_title.authors.all()
                            
                            for name in author_names:
                                # Case-insensitive author lookup
                                existing_author = Author.objects.filter(author_name__iexact=name).first()
                                if existing_author:
                                    author = existing_author
                                else:
                                    author = Author.objects.create(author_name=name)
                                
                                if author not in current_authors:
                                    AuthorDetail.objects.create(author=author, book_title=book_title)
                            
                            # 5. Smart duplicate detection for Excel import
                            # Also use case-insensitive publisher matching
                            existing_books = Book.objects.filter(
                                book_title=book_title,
                                publish_year=publish_year,
                                publisher__iexact=publisher
                            )
                            
                            # Check ISBN and Edition
                            book = None
                            if isbn:
                                existing_books = existing_books.filter(isbn=isbn)
                            
                            # Different edition = different book
                            if edition:
                                existing_books = existing_books.filter(edition=edition)
                            else:
                                existing_books = existing_books.filter(edition__isnull=True) | existing_books.filter(edition='')
                            
                            if existing_books.exists():
                                # Book exists with same details → update quantity
                                book = existing_books.first()
                            else:
                                # Create new Book record for different version
                                # Set quantity=1 initially (min validator requirement)
                                # BookImportDetail.save() will handle the actual quantity update
                                book = Book.objects.create(
                                    book_title=book_title,
                                    publish_year=publish_year,
                                    publisher=publisher,
                                    isbn=isbn if isbn else None,
                                    quantity=1,
                                    remaining_quantity=1,
                                    unit_price=unit_price,
                                    edition=edition,
                                    language=language
                                )
                                # Mark as just created so we know to set quantity properly
                                book._just_created = True
                            
                            # Note: BookImportDetail.save() handles quantity increment
                            # and BookItem creation automatically
                            
                            
                            # 6. Create Import Detail
                            BookImportDetail.objects.create(
                                receipt=receipt,
                                book=book,
                                quantity=quantity,
                                unit_price=unit_price
                            )
                            
                            success_count += 1
                            
                        except Exception as e:
                            error_count += 1
                            errors.append(f"Dòng {index+2}: {str(e)}")
                
                messages.success(request, f"Đã nhập thành công {success_count} dòng. Lỗi {error_count} dòng.")
                if errors:
                    for err in errors[:5]: # Show first 5 errors
                        messages.warning(request, err)
                        
                return redirect('book_import_list')
                
            except ImportError:
                messages.error(request, "Hệ thống thiếu thư viện 'pandas' hoặc 'openpyxl'. Vui lòng liên hệ Admin.")
            except Exception as e:
                messages.error(request, f"Lỗi xử lý file: {str(e)}")
    else:
        form = BookImportExcelForm()
    
    return render(request, 'app/books/book_import_excel.html', {
        'form': form,
        'page_title': 'Nhập sách từ Excel'
    })


@permission_required('Lập phiếu nhập sách', 'view')
def book_import_detail_view(request, import_id):
    """
    Xem chi tiết phiếu nhập sách
    """
    receipt = get_object_or_404(BookImportReceipt, id=import_id)
    import_details = receipt.import_details.all()
    
    # Check permission for cancel action
    from .decorators import check_permission
    can_cancel_import = check_permission(request.user, 'Lập phiếu nhập sách', 'delete')
    
    context = {
        'receipt': receipt,
        'import_details': import_details,
        'can_cancel_import': can_cancel_import,
        'page_title': f'Phiếu nhập sách #{receipt.id}'
    }
    
    return render(request, 'app/books/book_import_detail.html', context)


@permission_required('Lập phiếu nhập sách', 'delete')
def book_import_cancel_view(request, import_id):
    """
    Hủy phiếu nhập sách và XÓA sách đã nhập khỏi CSDL
    
    Business Logic:
    - Kiểm tra thời hạn hủy (từ tham số cancellation_time_limit)
    - Kiểm tra không có sách nào đang được mượn
    - Xóa BookItem được tạo từ phiếu nhập này
    - Giảm quantity của Book tương ứng
    - Đánh dấu phiếu đã hủy với audit trail
    """
    receipt = get_object_or_404(BookImportReceipt, id=import_id)
    
    # Kiểm tra phiếu đã bị hủy chưa
    if receipt.is_cancelled:
        messages.error(request, f'Phiếu nhập #{receipt.id} đã được hủy trước đó.')
        return redirect('book_import_detail', import_id=receipt.id)
    
    # Kiểm tra thời hạn hủy (sử dụng tham số từ CSDL)
    params = Parameter.objects.first()
    cancellation_hours = params.cancellation_time_limit if params else 24
    
    time_since_import = timezone.now() - receipt.import_date
    if time_since_import.total_seconds() > cancellation_hours * 3600:
        messages.error(
            request, 
            f'Không thể hủy phiếu nhập #{receipt.id}. '
            f'Chỉ được hủy trong vòng {cancellation_hours} giờ kể từ khi nhập.'
        )
        return redirect('book_import_detail', import_id=receipt.id)
    
    # Kiểm tra sách có đang được mượn không
    borrowed_books = []
    for detail in receipt.import_details.all():
        # Kiểm tra nếu có bất kỳ BookItem nào của sách này đang được mượn
        borrowed_count = BookItem.objects.filter(
            book=detail.book,
            is_borrowed=True
        ).count()
        
        if borrowed_count > 0:
            borrowed_books.append(f"{detail.book.book_title.book_title} ({borrowed_count} cuốn)")
    
    if borrowed_books:
        messages.error(
            request,
            f'Không thể hủy phiếu nhập #{receipt.id}. '
            f'Các sách sau đang được mượn: {", ".join(borrowed_books)}. '
            f'Vui lòng đợi trả sách trước khi hủy phiếu.'
        )
        return redirect('book_import_detail', import_id=receipt.id)
    
    if request.method == 'POST':
        cancel_reason = request.POST.get('cancel_reason', '').strip()
        
        if not cancel_reason:
            messages.error(request, 'Vui lòng nhập lý do hủy phiếu.')
            context = {
                'receipt': receipt,
                'import_details': receipt.import_details.all(),
                'page_title': f'Hủy phiếu nhập #{receipt.id}'
            }
            return render(request, 'app/books/book_import_cancel_confirm.html', context)
        
        try:
            from django.db import transaction
            
            with transaction.atomic():
                # Track số liệu để báo cáo
                deleted_items_count = 0
                updated_books = []
                
                # Xử lý từng BookImportDetail
                for detail in receipt.import_details.all():
                    book = detail.book
                    quantity_to_remove = detail.quantity
                    
                    # Xóa BookItem (lấy N items KHÔNG ĐANG MƯỢN, mới nhất của book này)
                    # Sắp xếp theo id giảm dần để lấy những item mới tạo nhất
                    items_to_delete_ids = list(
                        BookItem.objects.filter(
                            book=book,
                            is_borrowed=False
                        ).order_by('-id')[:quantity_to_remove].values_list('id', flat=True)
                    )
                    
                    items_deleted = len(items_to_delete_ids)
                    
                    # Kiểm tra số lượng có đủ để xóa không
                    if items_deleted < quantity_to_remove:
                        raise Exception(
                            f'Không đủ sách để xóa cho "{book.book_title.book_title}". '
                            f'Cần xóa {quantity_to_remove} nhưng chỉ có {items_deleted} sách rảnh.'
                        )
                    
                    # Xóa các BookItem bằng cách sử dụng IDs
                    BookItem.objects.filter(id__in=items_to_delete_ids).delete()
                    deleted_items_count += items_deleted
                    
                    # Cập nhật số lượng Book
                    new_quantity = max(0, book.quantity - quantity_to_remove)
                    new_remaining = max(0, book.remaining_quantity - quantity_to_remove)
                    
                    # Sử dụng update() để bypass model validation
                    # (vì quantity có MinValueValidator(1) nhưng khi hủy phiếu có thể về 0)
                    Book.objects.filter(pk=book.pk).update(
                        quantity=new_quantity,
                        remaining_quantity=new_remaining
                    )
                    updated_books.append(f"{book.book_title.book_title} (-{quantity_to_remove} cuốn)")
                
                # Đánh dấu phiếu đã hủy với audit trail
                receipt.is_cancelled = True
                receipt.cancelled_at = timezone.now()
                receipt.cancelled_by = request.user
                receipt.cancel_reason = cancel_reason
                receipt.save(update_fields=[
                    'is_cancelled', 'cancelled_at', 'cancelled_by', 'cancel_reason'
                ])
            
            messages.success(
                request,
                f'Đã hủy phiếu nhập #{receipt.id}. '
                f'Đã xóa {deleted_items_count} cuốn sách khỏi kho. '
                f'Sách đã cập nhật: {", ".join(updated_books)}'
            )
            return redirect('book_import_list')
            
        except Exception as e:
            messages.error(request, f'Có lỗi xảy ra khi hủy phiếu: {str(e)}')
            return redirect('book_import_detail', import_id=receipt.id)
    
    # GET: Hiển thị trang xác nhận hủy
    context = {
        'receipt': receipt,
        'import_details': receipt.import_details.all(),
        'page_title': f'Hủy phiếu nhập #{receipt.id}',
        'cancellation_hours': params.cancellation_time_limit if params else 24,
    }
    
    return render(request, 'app/books/book_import_cancel_confirm.html', context)


@permission_required('Lập phiếu nhập sách', 'view')
def book_import_list_view(request):
    """
    Danh sách phiếu nhập sách
    """
    receipts = BookImportReceipt.objects.all().order_by('-import_date')
    
    # Filter theo tháng/năm nếu có
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    # Validate Inputs
    try:
        if year:
            year_int = int(year)
            import datetime
            current_year = datetime.datetime.now().year
            if not (2000 <= year_int <= current_year):
                messages.warning(request, f'Năm "{year}" không hợp lệ. Vui lòng nhập năm từ 2000 đến {current_year}.')
                year = None
    except ValueError:
        messages.warning(request, f'Năm "{year}" không hợp lệ. Vui lòng nhập số.')
        year = None

    try:
        if month:
            month_int = int(month)
            if not (1 <= month_int <= 12):
                month = None
    except ValueError:
        month = None
    
    if month and year:
        receipts = receipts.filter(
            import_date__month=month,
            import_date__year=year
        )
    elif year:
        receipts = receipts.filter(import_date__year=year)
    
    context = {
        'receipts': receipts,
        'page_title': 'Danh sách phiếu nhập sách'
    }
    
    return render(request, 'app/books/book_import_list.html', context)

@permission_required('Lập phiếu nhập sách', 'view')
def download_book_import_template(request):
    """
    Tải file Excel mẫu để nhập sách.
    Nếu tồn tại /docs/mau_nhap.xlsx thì dùng file đó.
    Nếu không có thì tự tạo file mẫu và trả về.
    """
    import pandas as pd
    from django.http import HttpResponse, FileResponse
    from io import BytesIO
    from pathlib import Path

    # === Xác định đường dẫn file docs/mau_nhap.xlsx ===
    # views.py -> LibraryApp -> src -> project_root
    ROOT_DIR = Path("/")
    DOCS_DIR = ROOT_DIR / "docs"
    TEMPLATE_PATH = DOCS_DIR / "mau_nhap.xlsx"

    # === Nếu file tồn tại → trả file luôn ===
    if TEMPLATE_PATH.exists():
        return FileResponse(
            open(TEMPLATE_PATH, "rb"),
            as_attachment=True,
            filename="mau_nhap.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # === Không có file → tạo file mẫu ===
    sample_data = {
        'Tên sách': ['Đắc Nhân Tâm', 'Clean Code', 'Nhà Giả Kim'],
        'Thể loại': ['Tâm lý', 'Công nghệ', 'Văn học'],
        'Tác giả': ['Dale Carnegie', 'Robert C. Martin', 'Paulo Coelho'],
        'Năm XB': [2023, 2022, 2021],
        'NXB': ['NXB Trẻ', 'Prentice Hall', 'NXB Văn Học'],
        'Số lượng': [10, 5, 8],
        'Đơn giá': [80000, 350000, 95000],
        'Ngôn ngữ': ['Tiếng Việt', 'Tiếng Anh', 'Tiếng Việt'],
        'ISBN': ['978-604-1-12345-6', '978-0-13-235088-4', ''],
        'Phiên bản': ['Tái bản lần 5', 'First Edition', ''],
        'Mô tả': ['Sách về nghệ thuật đối nhân xử thế', 'Sách về lập trình sạch', '']
    }

    df = pd.DataFrame(sample_data)

    # Đảm bảo thư mục /docs tồn tại
    DOCS_DIR.mkdir(exist_ok=True)

    # Ghi file vào memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Nhập sách', index=False)

        worksheet = writer.sheets['Nhập sách']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(col)
            ) + 2
            worksheet.column_dimensions[chr(65 + idx)].width = max_length

    output.seek(0)

    # Đồng thời lưu file mẫu vào /docs để tái sử dụng
    with open(TEMPLATE_PATH, "wb") as f:
        f.write(output.getvalue())

    # Trả file cho người dùng
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="mau_nhap.xlsx"'
    return response


# ==================== BOOK SEARCH (YC3) ====================

def book_search_view(request):
    """
    Tra cứu sách - YC3
    Chức năng công khai - ai cũng có thể tra cứu
    """
    from django.db.models import Q
    
    form = BookSearchForm(request.GET)
    books = Book.objects.all().select_related('book_title', 'book_title__category').prefetch_related('book_title__authors')
    
    if form.is_valid():
        # Tìm kiếm theo tên sách hoặc mã sách
        search_text = form.cleaned_data.get('search_text')
        if search_text:
            books = books.filter(
                Q(book_title__book_title__icontains=search_text) |
                Q(id__icontains=search_text)
            )
        
        # Lọc theo thể loại
        category = form.cleaned_data.get('category')
        if category:
            books = books.filter(book_title__category=category)
        
        # Lọc theo tác giả
        author = form.cleaned_data.get('author')
        if author:
            books = books.filter(book_title__authors=author)
        
        # Lọc theo tình trạng
        status = form.cleaned_data.get('status')
        if status == 'available':
            books = books.filter(remaining_quantity__gt=0)
        elif status == 'unavailable':
            books = books.filter(remaining_quantity=0)
    
    # Sắp xếp theo tên sách
    books = books.order_by('book_title__book_title')
    
    # Phân trang (20 cuốn/trang)
    from django.core.paginator import Paginator
    paginator = Paginator(books, 20)
    page_num = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_num)
    
    context = {
        'form': form,
        'page_obj': page_obj,
        'books': page_obj.object_list,
        'total_results': paginator.count,
        'page_title': 'Tra cứu sách'
    }
    
    return render(request, 'app/books/book_search.html', context)


# ==================== BORROW BOOK (YC4) ====================

@permission_required('Lập phiếu mượn sách', 'add')
def borrow_book_view(request):
    """
    Cho mượn sách - YC4
    Validate theo thẻ còn hạn, không quá hạn, sách chưa mượn, không quá 5 quyển
    Hỗ trợ mượn nhiều sách cùng lúc
    """
    params = Parameter.objects.first()
    
    if request.method == 'POST':
        form = BorrowBookForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    reader_id = form.cleaned_data['reader_id']
                    book_ids = form.cleaned_data['book_ids']  # List of IDs
                    borrow_date = form.cleaned_data['borrow_date']
                    books = form.cleaned_data['books']  # Dict of Book objects
                    
                    # Lấy độc giả
                    reader = Reader.objects.get(id=reader_id)
                    
                    # Xử lý datetime từ form (đã là datetime, chỉ cần make_aware nếu cần)
                    from datetime import timedelta
                    borrow_datetime = borrow_date
                    if timezone.is_naive(borrow_datetime):
                        borrow_datetime = timezone.make_aware(borrow_datetime)
                    
                    # Tính ngày phải trả
                    due_date = borrow_datetime + timedelta(days=params.max_borrow_days)
                    
                    # Tạo phiếu mượn cho từng sách
                    receipts = []
                    for book_id in book_ids:
                        book = books[book_id]
                        
                        # Lấy cuốn sách còn sẵn (lock row to prevent race condition)
                        book_item = BookItem.objects.select_for_update().filter(book=book, is_borrowed=False).first()
                        if not book_item:
                            messages.error(request, f'Không tìm thấy cuốn "{book.book_title.book_title}" còn sẵn.')
                            return redirect('borrow_book')
                        
                        # Tạo phiếu mượn
                        borrow_receipt = BorrowReturnReceipt.objects.create(
                            reader=reader,
                            book_item=book_item,
                            borrow_date=borrow_datetime,
                            due_date=due_date,
                            notes=''
                        )
                        receipts.append(borrow_receipt)
                        
                        # Cập nhật trạng thái cuốn sách
                        book_item.is_borrowed = True
                        book_item.save(update_fields=['is_borrowed'])
                        
                        # Cập nhật số lượng còn lại của sách
                        book.remaining_quantity -= 1
                        book.save(update_fields=['remaining_quantity'])
                    
                    # Thông báo thành công
                    if len(receipts) == 1:
                        messages.success(
                            request,
                            f'Cho mượn thành công! Phiếu #{receipts[0].id} - Phải trả trước {due_date.strftime("%d/%m/%Y")}'
                        )
                        return redirect('borrow_book_detail', receipt_id=receipts[0].id)
                    else:
                        receipt_ids = ', '.join([f'#{r.id}' for r in receipts])
                        messages.success(
                            request,
                            f'Cho mượn {len(receipts)} quyển thành công! Phiếu: {receipt_ids} - Phải trả trước {due_date.strftime("%d/%m/%Y")}'
                        )
                        return redirect('borrow_book_list')
                    
            except Reader.DoesNotExist:
                messages.error(request, 'Độc giả không tồn tại.')
                return redirect('borrow_book')
            except Exception as e:
                messages.error(request, f'Lỗi: {str(e)}')
                return redirect('borrow_book')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    else:
        form = BorrowBookForm()
    
    context = {
        'form': form,
        'params': params,
        'page_title': 'Cho mượn sách'
    }
    
    return render(request, 'app/borrowing/borrow_book.html', context)


@permission_required('Quản lý mượn/trả', 'view')
def borrow_book_detail_view(request, receipt_id):
    """
    Xem chi tiết phiếu mượn sách
    """
    receipt = get_object_or_404(BorrowReturnReceipt, id=receipt_id)
    
    # Check permission for cancel action
    from .decorators import check_permission
    can_cancel_borrow = check_permission(request.user, 'Quản lý mượn/trả', 'delete')
    
    context = {
        'receipt': receipt,
        'can_cancel_borrow': can_cancel_borrow,
        'page_title': f'Chi tiết phiếu mượn sách #{receipt.id}'
    }
    
    return render(request, 'app/borrowing/borrow_book_detail.html', context)


@permission_required('Quản lý mượn/trả', 'delete')
def borrow_cancel_view(request, receipt_id):
    """
    Hủy phiếu mượn sách với audit trail
    
    Business Logic:
    - CHỈ hủy được phiếu CHƯA TRẢ (return_date is NULL)
    - Đánh dấu phiếu là đã hủy (is_cancelled=True)
    - Ghi lại: người hủy, thời gian hủy, lý do hủy
    - Rollback: Set book_item.is_borrowed = False, tăng book.remaining_quantity
    - Nếu có fine amount, trừ khỏi reader.total_debt
    """
    receipt = get_object_or_404(BorrowReturnReceipt, id=receipt_id)
    
    # Kiểm tra phiếu đã bị hủy chưa
    if receipt.is_cancelled:
        messages.error(request, f'Phiếu mượn #{receipt.id} đã được hủy trước đó.')
        return redirect('borrow_book_detail', receipt_id=receipt.id)
    
    # Kiểm tra đã trả sách chưa
    if receipt.return_date:
        messages.error(request, f'Không thể hủy phiếu mượn #{receipt.id} vì sách đã được trả. Đây là bản ghi lịch sử.')
        return redirect('borrow_book_detail', receipt_id=receipt.id)
    
    # Kiểm tra thời gian: chỉ hủy trong vòng N giờ kể từ khi mượn (N từ tham số hệ thống)
    params = Parameter.objects.first()
    cancellation_hours = params.cancellation_time_limit if params else 24
    
    time_since_borrow = timezone.now() - receipt.borrow_date
    if time_since_borrow.total_seconds() > cancellation_hours * 3600:
        messages.error(request, f'Không thể hủy phiếu mượn #{receipt.id}. Chỉ được hủy trong vòng {cancellation_hours} giờ kể từ khi mượn.')
        return redirect('borrow_book_detail', receipt_id=receipt.id)
    
    if request.method == 'POST':
        cancel_reason = request.POST.get('cancel_reason', '').strip()
        
        if not cancel_reason:
            messages.error(request, 'Vui lòng nhập lý do hủy phiếu.')
            context = {
                'receipt': receipt,
                'page_title': f'Hủy phiếu mượn #{receipt.id}'
            }
            return render(request, 'app/borrowing/borrow_cancel_confirm.html', context)
        
        try:
            # Đánh dấu phiếu đã hủy với audit trail
            receipt.is_cancelled = True
            receipt.cancelled_at = timezone.now()
            receipt.cancelled_by = request.user
            receipt.cancel_reason = cancel_reason
            receipt.save(update_fields=['is_cancelled', 'cancelled_at', 'cancelled_by', 'cancel_reason'])
            
            # Rollback: Un-borrow book
            receipt.book_item.is_borrowed = False
            receipt.book_item.save(update_fields=['is_borrowed'])
            
            receipt.book_item.book.remaining_quantity += 1
            receipt.book_item.book.save(update_fields=['remaining_quantity'])
            
            # Nếu có fine, trừ khỏi reader debt
            if receipt.fine_amount > 0:
                receipt.reader.total_debt -= receipt.fine_amount
                receipt.reader.save(update_fields=['total_debt'])
            
            messages.success(
                request,
                f'Đã hủy phiếu mượn #{receipt.id}. Sách đã được trả lại kho.'
            )
            return redirect('borrow_book_list')
            
        except Exception as e:
            messages.error(request, f'Có lỗi xảy ra khi hủy phiếu: {str(e)}')
            return redirect('borrow_book_detail', receipt_id=receipt.id)
    
    # GET: Hiển thị trang xác nhận hủy
    context = {
        'receipt': receipt,
        'page_title': f'Hủy phiếu mượn #{receipt.id}'
    }
    
    return render(request, 'app/borrowing/borrow_cancel_confirm.html', context)


@permission_required('Quản lý mượn/trả', 'view')
def borrow_book_list_view(request):
    """
    Danh sách phiếu mượn sách
    """
    from django.db.models import Q
    from django.utils import timezone
    
    # Mặc định sort
    receipts = BorrowReturnReceipt.objects.all().order_by('-borrow_date')
    
    # Filter theo trạng thái (nhận cả 'filter' và 'status' param)
    status = request.GET.get('filter') or request.GET.get('status', 'all')
    
    if status == 'unreturned':
        receipts = receipts.filter(return_date__isnull=True)
    elif status == 'overdue':
        receipts = receipts.filter(return_date__isnull=True, due_date__lt=timezone.now())
    elif status == 'returned':
        receipts = receipts.filter(return_date__isnull=False)
    # 'all' = no filter, show all
    
    # Search
    search = request.GET.get('search', '')
    if search:
        receipts = receipts.filter(
            Q(reader__reader_name__icontains=search) |
            Q(book_item__book__book_title__book_title__icontains=search) |
            Q(reader__email__icontains=search) |
            Q(id__icontains=search)
        )
    
    # Filter theo độc giả (giữ lại logic cũ phòng khi dùng)
    reader_id = request.GET.get('reader_id')
    if reader_id:
        receipts = receipts.filter(reader_id=reader_id)
        
    # Phân trang
    from django.core.paginator import Paginator
    paginator = Paginator(receipts, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'receipts': page_obj.object_list,
        'current_status': status,
        'search': search,
        'total_results': paginator.count,
        'page_title': 'Danh sách phiếu mượn sách'
    }
    
    return render(request, 'app/borrowing/borrow_book_list.html', context)


# ==================== API ENDPOINTS FOR BORROW BOOK ====================

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q

@permission_required('Quản lý độc giả', 'view')
@require_http_methods(["GET"])
def api_readers_list(request):
    """
    API lấy danh sách độc giả với tìm kiếm
    Query params: search (tìm kiếm theo tên hoặc email)
    """
    search = request.GET.get('search', '').strip()
    
    # Chỉ hiện độc giả hoạt động
    readers = Reader.objects.filter(is_active=True).order_by('reader_name')
    
    if search:
        readers = readers.filter(
            Q(reader_name__icontains=search) | 
            Q(email__icontains=search)
        )
    
    data = [
        {
            'id': reader.id,
            'name': reader.reader_name,
            'email': reader.email,
            'display': f"{reader.reader_name} - {reader.email}"
        }
        for reader in readers[:50]  # Giới hạn 50 kết quả
    ]
    
    return JsonResponse({'success': True, 'data': data})


@permission_required('Quản lý kho sách', 'view')
@require_http_methods(["GET"])
def api_books_list(request):
    """
    API lấy danh sách sách còn sẵn với tìm kiếm
    Query params: search (tìm kiếm theo tên sách hoặc tác giả)
    """
    search = request.GET.get('search', '').strip()
    
    # Chỉ hiện sách còn sẵn
    books = Book.objects.filter(remaining_quantity__gt=0).select_related(
        'book_title', 'book_title__category'
    ).order_by('book_title__book_title')
    
    if search:
        books = books.filter(
            Q(book_title__book_title__icontains=search) |
            Q(book_title__authors__author_name__icontains=search)
        ).distinct()
    
    data = [
        {
            'id': book.id,
            'title': book.book_title.book_title,
            'year': book.publish_year,
            'category': book.book_title.category.category_name if book.book_title.category else 'N/A',
            'remaining': book.remaining_quantity,
            'display': f"{book.book_title.book_title} ({book.publish_year}) - Còn {book.remaining_quantity} quyển"
        }
        for book in books[:50]  # Giới hạn 50 kết quả
    ]
    
    return JsonResponse({'success': True, 'data': data})


@permission_required('Quản lý mượn/trả', 'view')
@require_http_methods(["GET"])
def api_borrowing_readers(request):
    """
    API lấy danh sách độc giả đang mượn sách (chưa trả)
    """
    # Lấy tất cả phiếu mượn chưa trả, group by reader
    from django.db.models import Count, Max
    
    borrowing_records = BorrowReturnReceipt.objects.filter(
        return_date__isnull=True
    ).values('reader_id').annotate(
        count=Count('id'),
        latest_due=Max('due_date')
    ).order_by('-latest_due')
    
    records = list(borrowing_records[:100])  # Giới hạn 100 kết quả
    reader_ids = [r['reader_id'] for r in records]
    readers_map = Reader.objects.filter(id__in=reader_ids).in_bulk()
    
    data = []
    for record in records:
        if record['reader_id'] not in readers_map:
            continue
            
        reader = readers_map[record['reader_id']]
        
        # Lấy danh sách sách đang mượn
        borrows = BorrowReturnReceipt.objects.filter(
            reader_id=reader.id,
            return_date__isnull=True
        ).select_related('book_item__book__book_title')
        
        # Kiểm tra có quá hạn không
        from django.utils import timezone
        today = timezone.now().date()
        is_overdue = any(b.due_date.date() < today for b in borrows)
        
        data.append({
            'reader_id': reader.id,
            'reader_name': reader.reader_name,
            'reader_email': reader.email,
            'borrowed_count': record['count'],
            'latest_due_date': record['latest_due'].strftime('%d/%m/%Y'),
            'is_overdue': is_overdue,
            'books': [
                {
                    'id': b.book_item.book.id,
                    'title': b.book_item.book.book_title.book_title,
                    'borrow_date': b.borrow_date.strftime('%d/%m/%Y'),
                    'due_date': b.due_date.strftime('%d/%m/%Y'),
                    'is_overdue': b.due_date.date() < today
                }
                for b in borrows
            ]
        })
    
    return JsonResponse({'success': True, 'data': data})
"""
Views cho YC5: Nhận trả sách
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from django.utils import timezone
from django.db import models
from .models import BorrowReturnReceipt, Parameter, Reader
from .forms import ReturnBookForm


@permission_required('Lập phiếu trả sách', 'add')
def return_book_view(request):
    """
    Trang nhận trả sách - YC5
    Chọn độc giả → chọn sách → nhập ngày trả
    """
    params = Parameter.objects.first()
    fine_rate = params.fine_rate if params else 1000
    
    context = {
        'page_title': 'Nhận trả sách',
        'params': params,
        'fine_rate': fine_rate,
    }
    
    if request.method == 'POST':
        form = ReturnBookForm(request.POST)
        if form.is_valid():
            try:
                receipts = form.save()
            except ValidationError as e:
                # Bắt lỗi validation từ model (ví dụ: ngày trả < ngày mượn)
                for field, errors in e.message_dict.items():
                    for error in errors:
                        messages.error(request, error)
                return redirect('return_book')
            if receipts:
                count = len(receipts)
                # Lấy reader từ receipt đầu tiên và REFRESH từ DB để có total_debt mới nhất
                reader_id = receipts[0].reader_id if receipts else None
                if reader_id:
                    reader = Reader.objects.get(id=reader_id)
                    if reader.total_debt > 0:
                        messages.success(
                            request, 
                            f'Đã ghi nhận trả sách - {count} quyển. '
                            f'Độc giả còn nợ {reader.total_debt:,}đ. '
                            f'<a href="/receipt/" class="text-blue-600 dark:text-blue-50 underline font-semibold">Lập phiếu thu ngay</a>',
                            extra_tags='safe'
                        )
                    else:
                        messages.success(request, f'Đã ghi nhận trả sách - {count} quyển')
                else:
                    messages.success(request, f'Đã ghi nhận trả sách - {count} quyển')
                return redirect('return_book_list')
            else:
                messages.error(request, 'Lỗi khi lưu thông tin trả sách')
    else:
        form = ReturnBookForm()
    
    # Convert params to JSON for JavaScript
    import json
    params_json = json.dumps({
        'fine_rate': params.fine_rate if params else 1000
    })
    
    # Get all readers for dropdown
    readers = Reader.objects.all().values('id', 'reader_name', 'email').order_by('reader_name')
    
    context['form'] = form
    context['params_json'] = params_json
    context['readers'] = list(readers)
    context['readers_json'] = json.dumps(list(readers))
    return render(request, 'app/borrowing/return_book.html', context)


@permission_required('Lập phiếu trả sách', 'view')
def return_book_detail_view(request, receipt_id):
    """
    Xem chi tiết phiếu trả sách - YC5
    Hiển thị: thông tin độc giả, sách trả, tiền phạt
    """
    receipt = get_object_or_404(BorrowReturnReceipt, id=receipt_id)
    params = Parameter.objects.first()
    fine_rate = params.fine_rate if params else 1000
    
    # Tính tiền phạt
    days_overdue = receipt.days_overdue if receipt.is_overdue else 0
    fine_amount = days_overdue * fine_rate
    
    # Check permission for cancel return action
    from .decorators import check_permission
    can_cancel_return = check_permission(request.user, 'Quản lý mượn/trả', 'delete')
    
    context = {
        'page_title': f'Chi tiết phiếu trả sách #{receipt_id}',
        'receipt': receipt,
        'days_overdue': days_overdue,
        'fine_amount': fine_amount,
        'fine_rate': fine_rate,
        'can_cancel_return': can_cancel_return,
    }
    
    return render(request, 'app/borrowing/return_book_detail.html', context)


@permission_required('Quản lý mượn/trả', 'delete')
def return_cancel_view(request, receipt_id):
    """
    Hủy hành động trả sách (reverse return)
    
    Business Logic:
    - CHỈ hủy được phiếu ĐÃ TRẢ (return_date is NOT NULL)
    - Đặt lại return_date = NULL
    - Đánh dấu sách là đang mượn (is_borrowed = True)
    - Giảm remaining_quantity
    - Nếu đã thu tiền phạt (fine_amount > 0), hoàn lại vào reader.total_debt
    - Ghi audit trail
    """
    receipt = get_object_or_404(BorrowReturnReceipt, id=receipt_id)
    
    # Kiểm tra phiếu đã bị hủy chưa
    if receipt.is_cancelled:
        messages.error(request, f'Phiếu mượn #{receipt.id} đã được hủy trước đó. Không thể hủy hành động trả.')
        return redirect('return_book_detail', receipt_id=receipt.id)
    
    # Kiểm tra phải là phiếu đã trả
    if not receipt.return_date:
        messages.error(request, f'Phiếu #{receipt.id} chưa trả sách. Không có gì để hủy.')
        return redirect('return_book_detail', receipt_id=receipt.id)
    
    # Kiểm tra thời gian: chỉ hủy trong vòng N giờ kể từ khi trả (N từ tham số hệ thống)
    params = Parameter.objects.first()
    cancellation_hours = params.cancellation_time_limit if params else 24
    
    time_since_return = timezone.now() - receipt.return_date
    if time_since_return.total_seconds() > cancellation_hours * 3600:
        messages.error(request, f'Không thể hủy hành động trả sách #{receipt.id}. Chỉ được hủy trong vòng {cancellation_hours} giờ kể từ khi trả.')
        return redirect('return_book_detail', receipt_id=receipt.id)
    
    if request.method == 'POST':
        cancel_reason = request.POST.get('cancel_reason', '').strip()
        
        if not cancel_reason:
            messages.error(request, 'Vui lòng nhập lý do hủy hành động trả sách.')
            context = {
                'receipt': receipt,
                'page_title': f'Hủy hành động trả sách #{receipt.id}'
            }
            return render(request, 'app/borrowing/return_cancel_confirm.html', context)
        
        try:
            # Lưu thông tin cũ
            old_return_date = receipt.return_date
            old_fine = receipt.fine_amount
            
            # Đặt lại return_date = NULL (reverse return action)
            receipt.return_date = None
            
            # LƯU Ý: KHÔNG set is_cancelled = True vì:
            # - Đây là REVERSE RETURN, không phải HỦY PHIẾU
            # - Phiếu mượn vẫn VALID, chỉ quay lại trạng thái "đang mượn"
            # - Nếu set is_cancelled = True → home_view sẽ filter ra → bug count!
            
            # Ghi lý do trong notes (không dùng cancel_reason)
            if not receipt.notes:
                receipt.notes = f"[{timezone.now().strftime('%d/%m/%Y %H:%M')}] Đã hủy hành động trả sách (trả lúc {old_return_date.strftime('%d/%m/%Y %H:%M')}): {cancel_reason}"
            else:
                receipt.notes += f"\n[{timezone.now().strftime('%d/%m/%Y %H:%M')}] Đã hủy hành động trả sách (trả lúc {old_return_date.strftime('%d/%m/%Y %H:%M')}): {cancel_reason}"
            
            receipt.save(update_fields=['return_date', 'notes'])
            
            # Đánh dấu sách lại là đang mượn
            receipt.book_item.is_borrowed = True
            receipt.book_item.save(update_fields=['is_borrowed'])
            
            # Giảm remaining_quantity
            receipt.book_item.book.remaining_quantity -= 1
            receipt.book_item.book.save(update_fields=['remaining_quantity'])
            
            # HOÀN TIỀN PHẠT: Trừ fine_amount khỏi total_debt
            # - Fine đã được cộng vào total_debt khi trả sách (BorrowReturnReceipt.save())
            # - Khi hủy return, cần trừ fine_amount khỏi total_debt để nợ quay về trạng thái "dự tính"
            # - Nợ dự tính sẽ được tự động tính lại qua pending_debt property
            # - Phiếu quay lại trạng thái "đang mượn" nên fine sẽ được tính lại khi trả lần sau
            if old_fine > 0:
                receipt.reader.total_debt -= old_fine
                receipt.reader.save(update_fields=['total_debt'])
            
            messages.success(
                request,
                f'Đã hủy hành động trả sách #{receipt.id}. Sách được đánh dấu lại là đang mượn. '
                f'Tiền phạt {old_fine:,}đ đã được chuyển về trạng thái dự tính.'
            )
            return redirect('borrow_book_list')
            
        except Exception as e:
            messages.error(request, f'Có lỗi xảy ra khi hủy: {str(e)}')
            return redirect('return_book_detail', receipt_id=receipt.id)
    
    # GET: Hiển thị trang xác nhận
    context = {
        'receipt': receipt,
        'page_title': f'Hủy hành động trả sách #{receipt.id}'
    }
    
    return render(request, 'app/borrowing/return_cancel_confirm.html', context)


@permission_required('Lập phiếu trả sách', 'view')
def return_book_list_view(request):
    """
    Danh sách phiếu trả sách - YC5
    Hỗ trợ lọc và tìm kiếm
    """
    # Lấy bộ lọc từ GET
    filter_type = request.GET.get('filter', 'all')
    search = request.GET.get('search', '')
    
    # Base query: chỉ lấy phiếu đã trả (return_date != null)
    receipts = BorrowReturnReceipt.objects.filter(
        return_date__isnull=False
    ).select_related('reader', 'book_item__book__book_title')
    
    # Lọc theo loại
    # Lọc theo loại
    if filter_type != 'all':
        from django.db.models import F as DjangoF
        
        if filter_type == 'overdue':
             # Quá hạn: ngày trả > ngày phải trả
            receipts = receipts.filter(return_date__gt=DjangoF('due_date'))
        elif filter_type in ['ontime', 'returned']:
             # Đúng hạn: ngày trả <= ngày phải trả
            receipts = receipts.filter(return_date__lte=DjangoF('due_date'))
    
    # Tìm kiếm theo tên độc giả hoặc tên sách
    if search:
        from django.db.models import Q
        receipts = receipts.filter(
            Q(reader__reader_name__icontains=search) |
            Q(book_item__book__book_title__book_title__icontains=search) |
            Q(reader__email__icontains=search)
        )
    
    # Sắp xếp
    receipts = receipts.order_by('-return_date')
    
    # Phân trang
    from django.core.paginator import Paginator
    paginator = Paginator(receipts, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    params = Parameter.objects.first()
    fine_rate = params.fine_rate if params else 1000
    # Tính tiền phạt cho mỗi receipt
    for receipt in page_obj.object_list:
        if hasattr(receipt, 'is_overdue') and receipt.is_overdue:
            receipt.fine_amount = receipt.days_overdue * fine_rate
        else:
            receipt.fine_amount = 0
    
    context = {
        'page_title': 'Danh sách phiếu trả sách',
        'page_obj': page_obj,
        'receipts': page_obj.object_list,
        'filter_type': filter_type,
        'search': search,
        'total_results': paginator.count,
    }
    
    return render(request, 'app/borrowing/return_book_list.html', context)


@login_required
@permission_required('Lập phiếu trả sách', 'view')
@require_http_methods(["GET"])
def api_unreturned_receipts(request):
    """
    API: Lấy danh sách phiếu mượn chưa trả
    Dùng cho dropdown chọn phiếu trả sách
    """
    search = request.GET.get('search', '')
    print(f'[API] Fetching unreturned receipts, search={search}')
    
    # Lấy phiếu chưa trả
    receipts = BorrowReturnReceipt.objects.filter(
        return_date__isnull=True
    ).select_related('reader', 'book_item__book__book_title')
    
    print(f'[API] Found {len(receipts)} unreturned receipts')
    
    # Tìm kiếm
    if search:
        from django.db.models import Q
        receipts = receipts.filter(
            Q(reader__reader_name__icontains=search) |
            Q(book_item__book__book_title__book_title__icontains=search) |
            Q(reader__email__icontains=search) |
            Q(id__icontains=search)
        )
    
    # Giới hạn kết quả
    receipts = receipts[:50]
    
    data = []
    for receipt in receipts:
        try:
            days_borrowed = (timezone.now().date() - receipt.borrow_date.date()).days
            book_title = receipt.book_item.book.book_title.book_title
            reader_name = receipt.reader.reader_name
            
            data.append({
                'id': receipt.id,
                'reader_name': reader_name,
                'reader_email': receipt.reader.email,
                'book_title': book_title,
                'borrow_date': receipt.borrow_date.strftime('%d/%m/%Y'),
                'due_date': receipt.due_date.strftime('%d/%m/%Y'),
                'days_borrowed': days_borrowed,
                'is_overdue': receipt.due_date.date() < timezone.now().date(),
                'display': f"#{receipt.id} - {reader_name} ({book_title})"
            })
        except Exception as e:
            print(f'Error processing receipt {receipt.id}: {e}')
            continue
    
    return JsonResponse({'success': True, 'data': data})


@login_required
@permission_required('Quản lý mượn/trả', 'view')
@require_http_methods(["GET"])
def api_reader_borrowed_books(request, reader_id):
    """
    API: Lấy danh sách sách độc giả đó mượn nhưng chưa trả
    """
    try:
        reader = Reader.objects.get(id=reader_id)
    except Reader.DoesNotExist:
        return JsonResponse({'success': False, 'data': [], 'error': 'Reader not found'}, status=404)
    
    # Lấy danh sách phiếu mượn chưa trả của độc giả (loại trừ phiếu đã hủy)
    receipts = BorrowReturnReceipt.objects.filter(
        reader=reader,
        return_date__isnull=True,
        is_cancelled=False  # Loại trừ phiếu đã hủy
    ).select_related('book_item__book__book_title')
    
    data = []
    for receipt in receipts:
        try:
            today = timezone.localdate()
            days_borrowed = (today - receipt.borrow_date.date()).days
            book_title = receipt.book_item.book.book_title.book_title
            
            # Tính days_overdue từ due_date thực sự, converted sang localtime
            due_date = receipt.due_date
            if due_date:
                due_date = timezone.localtime(due_date).date()
            else:
                due_date = timezone.localtime(receipt.borrow_date).date()
            
            days_overdue = max(0, (today - due_date).days)
            
            data.append({
                'receipt_id': receipt.id,
                'book_item_id': receipt.book_item.id,
                'book_title': book_title,
                'barcode': receipt.book_item.barcode,
                'borrow_date': receipt.borrow_date.strftime('%d/%m/%Y'),
                'due_date': receipt.due_date.strftime('%d/%m/%Y'),
                'days_borrowed': days_borrowed,
                'days_overdue': days_overdue,
                'is_overdue': days_overdue > 0,
            })
        except Exception as e:
            print(f'Error processing receipt {receipt.id}: {e}')
            continue
    
    return JsonResponse({'success': True, 'data': data})


# ==================== PAYMENT MANAGEMENT (YC6) ====================

@permission_required('Lập phiếu thu tiền phạt', 'add')
def receipt_form_view(request):
    """
    Lập phiếu thu tiền phạt - YC6 (BM6)
    GET: Hiển thị form chọn độc giả → hiển thị nợ → nhập số tiền → submit
    POST: Xử lý submit, kiểm tra QĐ6, lưu phiếu, trừ nợ
    """
    import json
    
    # Lấy danh sách độc giả có nợ
    readers_with_debt = Reader.objects.filter(
        total_debt__gt=0
    ).values('id', 'reader_name', 'email', 'total_debt')
    
    readers_json = json.dumps(list(readers_with_debt))
    
    # Lấy tham số hệ thống
    try:
        params = Parameter.objects.first()
    except:
        params = None
    
    if request.method == 'POST':
        # Kiểm tra quyền 'add' cho lập phiếu thu
        from .decorators import check_permission
        if not check_permission(request.user, 'Lập phiếu thu tiền phạt', 'add'):
            messages.error(request, 'Bạn không có quyền thêm phiếu thu tiền.')
            return redirect('receipt_list')
        
        form = ReceiptForm(request.POST)
        if form.is_valid():
            receipt = form.save()
            if receipt:
                messages.success(
                    request,
                    f'Đã ghi nhận thu tiền {receipt.collected_amount:,}đ từ {receipt.reader.reader_name}. '
                    f'Nợ còn lại: {receipt.reader.total_debt:,}đ'
                )
                return redirect('receipt_list')
            else:
                messages.error(request, 'Lỗi khi lưu phiếu thu tiền')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ReceiptForm()
    
    # Bank configuration for VietQR - Query from database
    bank_account = BankAccount.objects.filter(is_active=True).first()
    
    if bank_account:
        bank_config = {
            'bankId': bank_account.bank_id,
            'accountNo': bank_account.account_no,
            'accountName': bank_account.account_name,
            'template': bank_account.template
        }
    else:
        # Fallback nếu chưa có tài khoản trong DB
        bank_config = {
            'bankId': '',
            'accountNo': '',
            'accountName': '',
            'template': 'print'
        }
    
    context = {
        'form': form,
        'readers_json': readers_json,
        'params': params,
        'bank_config': json.dumps(bank_config),
    }
    
    return render(request, 'app/receipts/receipt_form.html', context)


@permission_required('Quản lý phiếu thu', 'view')
def receipt_list_view(request):
    """
    Danh sách phiếu thu tiền
    """
    from .decorators import check_permission
    from django.db.models import Q
    from django.core.paginator import Paginator
    
    # Get all receipts (including cancelled)
    receipts = Receipt.objects.select_related('reader').all().order_by('-created_date')
    
    # Status filter
    status = request.GET.get('status', '')
    if status == 'active':
        receipts = receipts.filter(is_cancelled=False)
    elif status == 'cancelled':
        receipts = receipts.filter(is_cancelled=True)
    # else: show all
    
    # Search by reader name or email
    search = request.GET.get('search', '')
    if search:
        receipts = receipts.filter(
            Q(reader__reader_name__icontains=search) | 
            Q(reader__email__icontains=search)
        )
    
    # Date range filter
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date:
        receipts = receipts.filter(created_date__gte=from_date)
    if to_date:
        # Include the entire end date
        to_date_obj = timezone.datetime.strptime(to_date, '%Y-%m-%d')
        to_date_end = to_date_obj.replace(hour=23, minute=59, second=59)
        receipts = receipts.filter(created_date__lte=to_date_end)
    
    # Pagination
    paginator = Paginator(receipts, 20)  # 20 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'receipts': page_obj,
        'page_obj': page_obj,
        'total_results': receipts.count(),
        'search': search,
        'from_date': from_date,
        'to_date': to_date,
        'status': status,
        'can_add_receipt': check_permission(request.user, 'Quản lý phiếu thu', 'add'),
        'page_title': 'Danh sách phiếu thu'
    }
    
    return render(request, 'app/receipts/receipt_list.html', context)


@permission_required('Quản lý phiếu thu', 'view')
def receipt_detail_view(request, receipt_id):
    """
    Chi tiết phiếu thu tiền
    """
    receipt = get_object_or_404(Receipt, id=receipt_id)
    
    # Check permission for cancel action
    from .decorators import check_permission
    can_cancel_receipt = check_permission(request.user, 'Quản lý phiếu thu', 'delete')
    
    context = {
        'receipt': receipt,
        'can_cancel_receipt': can_cancel_receipt,
        'page_title': f'Chi tiết phiếu thu #{receipt.id}'
    }
    
    return render(request, 'app/receipts/receipt_detail.html', context)


@permission_required('Quản lý phiếu thu', 'delete')
def receipt_cancel_view(request, receipt_id):
    """
    Hủy phiếu thu tiền với audit trail
    
    Business Logic:
    - Đánh dấu phiếu là đã hủy (is_cancelled=True)
    - Ghi lại: người hủy, thời gian hủy, lý do hủy
    - Hoàn tiền: cộng lại collected_amount vào reader.total_debt
    - Không cho phép hủy phiếu đã bị hủy trước đó
    """
    receipt = get_object_or_404(Receipt, id=receipt_id)
    
    # Kiểm tra phiếu đã bị hủy chưa
    if receipt.is_cancelled:
        messages.error(request, f'Phiếu thu #{receipt.id} đã được hủy trước đó.')
        return redirect('receipt_detail', receipt_id=receipt.id)
    
    # Kiểm tra thời gian: chỉ hủy trong vòng N giờ kể từ khi lập phiếu (N từ tham số hệ thống)
    params = Parameter.objects.first()
    cancellation_hours = params.cancellation_time_limit if params else 24
    
    time_since_created = timezone.now() - receipt.created_date
    if time_since_created.total_seconds() > cancellation_hours * 3600:
        messages.error(request, f'Không thể hủy phiếu thu #{receipt.id}. Chỉ được hủy trong vòng {cancellation_hours} giờ kể từ khi lập phiếu.')
        return redirect('receipt_detail', receipt_id=receipt.id)
    
    if request.method == 'POST':
        cancel_reason = request.POST.get('cancel_reason', '').strip()
        
        if not cancel_reason:
            messages.error(request, 'Vui lòng nhập lý do hủy phiếu.')
            context = {
                'receipt': receipt,
                'page_title': f'Hủy phiếu thu #{receipt.id}'
            }
            return render(request, 'app/receipts/receipt_cancel_confirm.html', context)
        
        try:
            # QUAN TRỌNG: Hoàn tiền cho độc giả TRƯỚC khi save receipt
            # Vì Receipt có validation check: collected_amount <= reader.total_debt
            receipt.reader.total_debt += receipt.collected_amount
            receipt.reader.save(update_fields=['total_debt'])
            
            # Sau đó mới đánh dấu phiếu đã hủy với audit trail
            receipt.is_cancelled = True
            receipt.cancelled_at = timezone.now()
            receipt.cancelled_by = request.user
            receipt.cancel_reason = cancel_reason
            receipt.save(update_fields=['is_cancelled', 'cancelled_at', 'cancelled_by', 'cancel_reason'])
            
            messages.success(
                request,
                f'Đã hủy phiếu thu #{receipt.id}. Đã hoàn {receipt.collected_amount:,}đ vào tổng nợ của {receipt.reader.reader_name}.'
            )
            return redirect('receipt_list')
            
        except Exception as e:
            messages.error(request, f'Có lỗi xảy ra khi hủy phiếu: {str(e)}')
            return redirect('receipt_detail', receipt_id=receipt.id)
    
    # GET: Hiển thị trang xác nhận hủy
    context = {
        'receipt': receipt,
        'page_title': f'Hủy phiếu thu #{receipt.id}'
    }
    
    return render(request, 'app/receipts/receipt_cancel_confirm.html', context)


@login_required
@permission_required('Quản lý độc giả', 'view')
@require_http_methods(["GET"])
def api_reader_debt(request, reader_id):
    """
    API: Lấy thông tin nợ tiền phạt của độc giả
    """
    try:
        reader = Reader.objects.get(id=reader_id)
        
        return JsonResponse({
            'success': True,
            'reader_id': reader.id,
            'reader_name': reader.reader_name,
            'email': reader.email,
            'total_debt': reader.total_debt,
        })
    except Reader.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Độc giả không tồn tại'
        }, status=404)


# ==================== REPORTS - YC7 ====================

@permission_required('Báo cáo mượn sách theo thể loại', 'view')
def report_borrow_by_category_view(request):
    """
    YC7 - BM7.1: Báo cáo thống kê tình hình mượn sách theo thể loại
    """
    # Lấy tháng/năm từ request, mặc định là tháng hiện tại
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    if not month or not year:
        now = timezone.now()
        month = now.month
        year = now.year
    else:
        month = int(month)
        year = int(year)
    
    # Tính ngày đầu và cuối tháng
    from calendar import monthrange
    first_day = timezone.datetime(year, month, 1, tzinfo=timezone.get_current_timezone())
    last_day_num = monthrange(year, month)[1]
    last_day = timezone.datetime(year, month, last_day_num, 23, 59, 59, tzinfo=timezone.get_current_timezone())
    
    # Lấy danh sách phiếu mượn trong tháng (D3)
    borrow_receipts = BorrowReturnReceipt.objects.filter(
        borrow_date__gte=first_day,
        borrow_date__lte=last_day
    ).select_related('book_item__book__book_title__category')
    
    # Đếm số lượt mượn theo thể loại (D4)
    category_stats = {}
    total_borrows = 0
    
    for receipt in borrow_receipts:
        # Lấy thể loại từ book_item -> book -> book_title -> category
        if receipt.book_item and receipt.book_item.book:
            category = receipt.book_item.book.book_title.category
            category_name = category.category_name
            
            if category_name not in category_stats:
                category_stats[category_name] = 0
            
            category_stats[category_name] += 1
            total_borrows += 1
    
    # Tính tỉ lệ (%)
    report_data = []
    for idx, (category_name, borrow_count) in enumerate(sorted(category_stats.items()), start=1):
        percentage = (borrow_count / total_borrows * 100) if total_borrows > 0 else 0
        report_data.append({
            'stt': idx,
            'category_name': category_name,
            'borrow_count': borrow_count,
            'percentage': round(percentage, 2)
        })
    
    context = {
        'report_data': report_data,
        'total_borrows': total_borrows,
        'month': month,
        'year': year,
        'page_title': f'Báo cáo mượn sách theo thể loại - Tháng {month}/{year}'
    }
    
    return render(request, 'app/reports/report_borrow_by_category.html', context)


@permission_required('Báo cáo sách trả trễ', 'view')
def report_overdue_books_view(request):
    """
    YC7 - BM7.2: Báo cáo thống kê sách trả trễ
    """
    # Lấy ngày báo cáo từ request, mặc định là ngày hiện tại
    report_date_str = request.GET.get('report_date')
    
    if report_date_str:
        try:
            report_date = timezone.datetime.strptime(report_date_str, '%Y-%m-%d')
            report_date = timezone.make_aware(report_date, timezone.get_current_timezone())
        except ValueError:
            report_date = timezone.now()
    else:
        report_date = timezone.now()
    
    # Lấy danh sách phiếu mượn đã trả trễ
    # return_date > due_date (đã trả nhưng trả trễ hơn hạn)
    overdue_receipts = BorrowReturnReceipt.objects.filter(
        return_date__isnull=False,  # Đã trả
        return_date__gt=models.F('due_date')  # Trả sau hạn
    ).select_related('book_item__book__book_title')
    
    # Tạo danh sách thống kê (D4)
    report_data = []
    for idx, receipt in enumerate(overdue_receipts, start=1):
        # Tính số ngày trễ = return_date - due_date
        overdue_days = (receipt.return_date.date() - receipt.due_date.date()).days
        
        book_title = receipt.book_item.book.book_title.book_title if receipt.book_item else "N/A"
        
        report_data.append({
            'stt': idx,
            'book_title': book_title,
            'borrow_date': receipt.borrow_date,
            'overdue_days': overdue_days
        })
    
    context = {
        'report_data': report_data,
        'report_date': report_date,
        'page_title': f'Báo cáo sách trả trễ - Ngày {report_date.strftime("%d/%m/%Y")}'
    }
    
    return render(request, 'app/reports/report_overdue_books.html', context)


@permission_required('Báo cáo mượn sách theo thể loại', 'view')
def report_borrow_by_category_excel(request):
    """
    Export báo cáo mượn sách theo thể loại ra file Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from django.http import HttpResponse
    from calendar import monthrange
    
    # Lấy tháng/năm từ request
    month = int(request.GET.get('month', timezone.now().month))
    year = int(request.GET.get('year', timezone.now().year))
    
    # Tính ngày đầu và cuối tháng
    first_day = timezone.datetime(year, month, 1, tzinfo=timezone.get_current_timezone())
    last_day_num = monthrange(year, month)[1]
    last_day = timezone.datetime(year, month, last_day_num, 23, 59, 59, tzinfo=timezone.get_current_timezone())
    
    # Lấy dữ liệu
    borrow_receipts = BorrowReturnReceipt.objects.filter(
        borrow_date__gte=first_day,
        borrow_date__lte=last_day
    )
    
    category_stats = {}
    total_borrows = 0
    
    for receipt in borrow_receipts:
        if receipt.book_item and receipt.book_item.book:
            category_name = receipt.book_item.book.book_title.category.category_name
            category_stats[category_name] = category_stats.get(category_name, 0) + 1
            total_borrows += 1
    
    # Tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Mượn sách T{month}-{year}"
    
    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = f'Báo cáo mượn sách theo thể loại - Tháng {month}/{year}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['STT', 'Tên thể loại', 'Số lượt mượn', 'Tỉ lệ (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 4
    for idx, (category_name, borrow_count) in enumerate(sorted(category_stats.items()), 1):
        percentage = round((borrow_count / total_borrows * 100), 2) if total_borrows > 0 else 0
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=category_name).border = thin_border
        ws.cell(row=row, column=3, value=borrow_count).border = thin_border
        ws.cell(row=row, column=4, value=f"{percentage}%").border = thin_border
        row += 1
    
    # Total
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value='Tổng số lượt mượn:').font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_borrows).font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="bao_cao_muon_sach_{month}_{year}.xlsx"'
    wb.save(response)
    return response


@permission_required('Báo cáo sách trả trễ', 'view')
def report_overdue_books_excel(request):
    """
    Export báo cáo sách trả trễ ra file Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from django.http import HttpResponse
    
    # Lấy ngày báo cáo
    report_date_str = request.GET.get('report_date')
    if report_date_str:
        try:
            report_date = timezone.datetime.strptime(report_date_str, '%Y-%m-%d')
            report_date = timezone.make_aware(report_date, timezone.get_current_timezone())
        except ValueError:
            report_date = timezone.now()
    else:
        report_date = timezone.now()
    
    # Lấy dữ liệu
    overdue_receipts = BorrowReturnReceipt.objects.filter(
        return_date__isnull=True,
        due_date__lt=report_date
    ).select_related('book_item__book__book_title')
    
    # Tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sách trả trễ"
    
    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = f'Báo cáo sách trả trễ - Ngày {report_date.strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['STT', 'Tên sách', 'Ngày mượn', 'Số ngày trễ']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 4
    for idx, receipt in enumerate(overdue_receipts, 1):
        overdue_days = (report_date.date() - receipt.due_date.date()).days
        book_title = receipt.book_item.book.book_title.book_title if receipt.book_item else "N/A"
        
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=book_title).border = thin_border
        ws.cell(row=row, column=3, value=receipt.borrow_date.strftime("%d/%m/%Y %H:%M")).border = thin_border
        ws.cell(row=row, column=4, value=overdue_days).border = thin_border
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="bao_cao_sach_tra_tre_{report_date.strftime("%Y_%m_%d")}.xlsx"'
    wb.save(response)
    return response


@permission_required('Báo cáo mượn sách theo thể loại', 'view')
def report_borrow_situation_view(request):
    """
    Báo cáo tình hình mượn sách theo khoảng thời gian
    """
    from datetime import timedelta
    
    # Lấy ngày từ request
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    
    # Mặc định: 1 năm trước đến hiện tại
    now = timezone.now()
    if from_date_str:
        try:
            from_date = timezone.datetime.strptime(from_date_str, '%Y-%m-%d')
            from_date = timezone.make_aware(from_date, timezone.get_current_timezone())
        except ValueError:
            from_date = now - timedelta(days=365)
    else:
        from_date = now - timedelta(days=365)
    
    if to_date_str:
        try:
            to_date = timezone.datetime.strptime(to_date_str, '%Y-%m-%d')
            to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59), timezone.get_current_timezone())
        except ValueError:
            to_date = now
    else:
        to_date = now
    
    # Lấy dữ liệu mượn trong khoảng thời gian
    borrow_receipts = BorrowReturnReceipt.objects.filter(
        borrow_date__gte=from_date,
        borrow_date__lte=to_date
    ).select_related('book_item__book__book_title__category')
    
    # Đếm theo thể loại
    category_stats = {}
    total_borrows = 0
    
    for receipt in borrow_receipts:
        if receipt.book_item and receipt.book_item.book:
            category_name = receipt.book_item.book.book_title.category.category_name
            category_stats[category_name] = category_stats.get(category_name, 0) + 1
            total_borrows += 1
    
    # Tạo dữ liệu báo cáo
    report_data = []
    for idx, (category_name, borrow_count) in enumerate(sorted(category_stats.items(), key=lambda x: -x[1]), 1):
        percentage = round((borrow_count / total_borrows * 100), 2) if total_borrows > 0 else 0
        report_data.append({
            'stt': idx,
            'category_name': category_name,
            'borrow_count': borrow_count,
            'percentage': percentage
        })
    
    # Dữ liệu cho biểu đồ (JSON)
    import json
    chart_labels = [item['category_name'] for item in report_data]
    chart_data = [item['borrow_count'] for item in report_data]
    
    context = {
        'report_data': report_data,
        'total_borrows': total_borrows,
        'from_date': from_date,
        'to_date': to_date,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        'page_title': 'Báo cáo Tình hình mượn sách'
    }
    
    return render(request, 'app/reports/report_borrow_situation.html', context)


@permission_required('Báo cáo', 'view')
def report_borrow_situation_excel(request):
    """
    Export báo cáo tình hình mượn sách ra file Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from django.http import HttpResponse
    from datetime import timedelta
    
    # Lấy ngày từ request
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    
    now = timezone.now()
    if from_date_str:
        try:
            from_date = timezone.datetime.strptime(from_date_str, '%Y-%m-%d')
            from_date = timezone.make_aware(from_date, timezone.get_current_timezone())
        except ValueError:
            from_date = now - timedelta(days=365)
    else:
        from_date = now - timedelta(days=365)
    
    if to_date_str:
        try:
            to_date = timezone.datetime.strptime(to_date_str, '%Y-%m-%d')
            to_date = timezone.make_aware(to_date.replace(hour=23, minute=59, second=59), timezone.get_current_timezone())
        except ValueError:
            to_date = now
    else:
        to_date = now
    
    # Lấy dữ liệu
    borrow_receipts = BorrowReturnReceipt.objects.filter(
        borrow_date__gte=from_date,
        borrow_date__lte=to_date
    ).select_related('book_item__book__book_title__category')
    
    category_stats = {}
    total_borrows = 0
    
    for receipt in borrow_receipts:
        if receipt.book_item and receipt.book_item.book:
            category_name = receipt.book_item.book.book_title.category.category_name
            category_stats[category_name] = category_stats.get(category_name, 0) + 1
            total_borrows += 1
    
    # Tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tình hình mượn sách"
    
    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = f'Báo cáo Tình hình mượn sách'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = f'Từ {from_date.strftime("%d/%m/%Y")} đến {to_date.strftime("%d/%m/%Y")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['STT', 'Tên thể loại', 'Số lượt mượn', 'Tỉ lệ (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 5
    for idx, (category_name, borrow_count) in enumerate(sorted(category_stats.items(), key=lambda x: -x[1]), 1):
        percentage = round((borrow_count / total_borrows * 100), 2) if total_borrows > 0 else 0
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=category_name).border = thin_border
        ws.cell(row=row, column=3, value=borrow_count).border = thin_border
        ws.cell(row=row, column=4, value=f"{percentage}%").border = thin_border
        row += 1
    
    # Total
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value='Tổng số lượt mượn:').font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_borrows).font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="bao_cao_tinh_hinh_muon_{from_date.strftime("%Y%m%d")}_{to_date.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


# ==================== SYSTEM PARAMETERS - YC8 ====================

@permission_required('Thay đổi quy định', 'view')
def parameter_update_view(request):
    """
    YC8 - Thay đổi quy định hệ thống
    
    Cho phép thay đổi:
    - Tuổi độc giả, thời hạn thẻ
    - Khoảng cách năm xuất bản
    - Số sách mượn tối đa, số ngày mượn
    - Đơn giá phạt
    - Quy định kiểm tra số tiền thu
    """
    # Lấy hoặc tạo bản ghi Parameter duy nhất
    parameter, created = Parameter.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        # Kiểm tra quyền edit cho POST request
        from .decorators import check_permission
        if not check_permission(request.user, 'Thay đổi quy định', 'edit'):
            messages.error(request, 'Bạn không có quyền sửa "Cài đặt hệ thống".')
            return redirect('parameter_update')
        
        form = ParameterForm(request.POST, instance=parameter)
        if form.is_valid():
            try:
                # Lưu các thay đổi (D4, D5)
                updated_param = form.save()
                
                messages.success(
                    request,
                    ' Cập nhật quy định thành công! '
                    # f'Tuổi: {updated_param.min_age}-{updated_param.max_age}, '
                    # f'Thời hạn thẻ: {updated_param.card_validity_period} tháng, '
                    # f'Số sách mượn tối đa: {updated_param.max_borrowed_books}, '
                    # f'Số ngày mượn tối đa: {updated_param.max_borrow_days}, '
                    # f'Tiền phạt: {updated_param.fine_rate:,}đ/ngày'
                )
            except Exception as e:
                messages.error(request, f'Có lỗi xảy ra: {str(e)}')
        else:
            # Hiển thị lỗi validation
            for field, errors in form.errors.items():
                # Lấy label thân thiện từ form field, bỏ qua '__all__' errors
                if field == '__all__':
                    # Non-field errors
                    for error in errors:
                        messages.error(request, str(error))
                else:
                    # Field-specific errors
                    field_label = form.fields[field].label if field in form.fields else field
                    for error in errors:
                        messages.error(request, f'{field_label}: {error}')
        
        # Always redirect to refresh form with database values
        return redirect('parameter_update')
    else:
        # GET: Hiển thị form với giá trị hiện tại (D3)
        form = ParameterForm(instance=parameter)
    
    # Lấy danh sách loại độc giả với số lượng
    reader_types = ReaderType.objects.annotate(
        reader_count=Count('readers')
    ).order_by('reader_type_name')
    
    context = {
        'form': form,
        'parameter': parameter,
        'reader_types': reader_types,
        'page_title': 'Thay đổi quy định hệ thống'
    }
    
    return render(request, 'app/settings/parameter_update.html', context)


# ==================== QUẢN LÝ NGƯỜI DÙNG ====================

@permission_required('Quản lý người dùng', 'view')
def user_list_view(request):
    """
    Danh sách người dùng - Chỉ dành cho Quản lý
    Hiển thị tất cả users với vai trò của họ
    """
    users = User.objects.all().order_by('-date_joined')
    
    # Thêm thông tin vai trò cho mỗi user
    user_data = []
    for user in users:
        role = 'Quản lý' if user.is_superuser else ('Thủ thư' if user.is_staff else 'Người dùng')
        user_data.append({
            'user': user,
            'role': role,
            'is_active_text': 'Đang hoạt động' if user.is_active else 'Đã khóa'
        })
    
    context = {
        'user_data': user_data,
        'page_title': 'Quản lý người dùng'
    }
    
    return render(request, 'app/users/user_list.html', context)


@permission_required('Quản lý người dùng', 'add')
def user_create_view(request):
    """
    Tạo người dùng mới - Chỉ dành cho Quản lý
    
    Cho phép tạo:
    - Thủ thư (is_staff=True)
    - Quản lý (is_superuser=True)
    """
    from .models import LibraryUser
    from datetime import date
    from django.contrib.auth.password_validation import validate_password
    from django.core.exceptions import ValidationError
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        user_group_id = request.POST.get('user_group')  # Vai trò = UserGroup
        
        # Validation
        if not username or not password:
            messages.error(request, 'Tên đăng nhập và mật khẩu là bắt buộc.')
        elif not user_group_id:
            messages.error(request, 'Vui lòng chọn vai trò.')
        elif password != password_confirm:
            messages.error(request, 'Mật khẩu xác nhận không khớp.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, f'Tên đăng nhập "{username}" đã tồn tại.')
        elif email and User.objects.filter(email=email).exists():
            messages.error(request, f'Email "{email}" đã được sử dụng.')
        else:
            try:
                # Validate password strength
                temp_user = User(username=username, email=email)
                validate_password(password, user=temp_user)
                
                with transaction.atomic():
                    # Lấy UserGroup đã chọn
                    user_group = UserGroup.objects.get(id=user_group_id)
                    role_name = user_group.user_group_name
                    
                    # Lấy thông tin mở rộng
                    phone_number = request.POST.get('phone_number', '').strip()
                    address = request.POST.get('address', '').strip()
                    position = request.POST.get('position', '').strip() or role_name
                    dob_str = request.POST.get('date_of_birth')
                    
                    try:
                        if dob_str:
                            date_of_birth = timezone.datetime.strptime(dob_str, '%Y-%m-%d').date()
                        else:
                            date_of_birth = date(1990, 1, 1)
                    except ValueError:
                        date_of_birth = date(1990, 1, 1)
                    
                    # Tạo user mới
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password,
                        first_name=first_name,
                        last_name=last_name
                    )
                    
                    # Gán vai trò dựa trên UserGroup
                    if role_name == 'Quản lý':
                        user.is_staff = True
                        user.is_superuser = True
                    else:
                        user.is_staff = True
                        user.is_superuser = False
                        
                    # Tạo LibraryUser liên kết với UserGroup
                    LibraryUser.objects.create(
                        user=user,
                        full_name=f"{first_name} {last_name}".strip() or username,
                        date_of_birth=date_of_birth,
                        position=position,
                        user_group=user_group,
                        phone_number=phone_number,
                        address=address,
                        email=email or '',
                        is_active=True
                    )
                    
                    user.save()
                    
                    messages.success(
                        request,
                        f'Đã tạo tài khoản "{username}" thành công!'
                    )
                    return redirect('user_list')
                    
            except ValidationError as e:
                error_msg = ' '.join(e.messages)
                messages.error(request, error_msg)
            except UserGroup.DoesNotExist:
                messages.error(request, 'Vai trò không hợp lệ.')
            except Exception as e:
                messages.error(request, f'Có lỗi xảy ra: {str(e)}')
    
    # Lấy danh sách user groups cho dropdown
    user_groups = UserGroup.objects.all().order_by('user_group_name')
    
    context = {
        'page_title': 'Tạo người dùng mới',
        'user_groups': user_groups
    }
    
    return render(request, 'app/users/user_form.html', context)


@permission_required('Quản lý người dùng', 'edit')
def user_edit_view(request, user_id):
    """
    Chỉnh sửa thông tin người dùng - Chỉ dành cho Quản lý
    
    Cho phép:
    - Thay đổi thông tin cơ bản
    - Thay đổi vai trò (UserGroup)
    - Khóa/Mở khóa tài khoản
    - Reset mật khẩu
    """
    from .models import LibraryUser
    user = get_object_or_404(User, id=user_id)
    
    # Lấy thông tin mở rộng LibraryUser (nếu có)
    try:
        library_user = LibraryUser.objects.get(user=user)
        current_group_id = library_user.user_group.id if library_user.user_group else None
    except LibraryUser.DoesNotExist:
        library_user = None
        current_group_id = None

    if request.method == 'POST':
        action = request.POST.get('action', 'update')
        
        if action == 'update':
            email = request.POST.get('email', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            user_group_id = request.POST.get('user_group')
            is_active = request.POST.get('is_active') == 'on'
            
            # Thông tin mở rộng
            phone_number = request.POST.get('phone_number', '').strip()
            address = request.POST.get('address', '').strip()
            position = request.POST.get('position', '').strip()
            dob_str = request.POST.get('date_of_birth')
            
            from django.utils import timezone
            from datetime import date
            try:
                if dob_str:
                    date_of_birth = timezone.datetime.strptime(dob_str, '%Y-%m-%d').date()
                else:
                    date_of_birth = date(1990, 1, 1)
            except ValueError:
                date_of_birth = date(1990, 1, 1)
            
            # Validation email
            if email and User.objects.filter(email=email).exclude(id=user_id).exists():
                messages.error(request, f'Email "{email}" đã được sử dụng bởi tài khoản khác.')
            else:
                try:
                    with transaction.atomic():
                        user.email = email
                        user.first_name = first_name
                        user.last_name = last_name
                        user.is_active = is_active
                        
                        # Cập nhật vai trò nếu có thay đổi
                        if user_group_id:
                            try:
                                user_group = UserGroup.objects.get(id=user_group_id)
                                role_name = user_group.user_group_name
                                
                                # Cập nhật User flags
                                if role_name == 'Quản lý':
                                    user.is_staff = True
                                    user.is_superuser = True
                                else:
                                    user.is_staff = True
                                    user.is_superuser = False
                                
                                # Cập nhật LibraryUser
                                if not library_user:
                                    library_user = LibraryUser.objects.create(
                                        user=user,
                                        full_name=f"{first_name} {last_name}".strip() or user.username,
                                        date_of_birth=date_of_birth,
                                        position=position or role_name,
                                        user_group=user_group,
                                        phone_number=phone_number,
                                        address=address,
                                        email=email or ''
                                    )
                                else:
                                    library_user.full_name = f"{first_name} {last_name}".strip() or user.username
                                    library_user.user_group = user_group
                                    library_user.position = position or role_name
                                    library_user.date_of_birth = date_of_birth
                                    library_user.phone_number = phone_number
                                    library_user.address = address
                                    library_user.is_active = is_active
                                    library_user.save()
                                    
                            except UserGroup.DoesNotExist:
                                pass # Giữ nguyên nếu ID không hợp lệ
                        else:
                            # Cập nhật thông tin LibraryUser kể cả khi không đổi role
                            if library_user:
                                library_user.full_name = f"{first_name} {last_name}".strip() or user.username
                                library_user.position = position if position else library_user.position
                                library_user.date_of_birth = date_of_birth
                                library_user.phone_number = phone_number
                                library_user.address = address
                                library_user.is_active = is_active
                                library_user.save()
                        
                        user.save()
                        
                        messages.success(request, f'Đã cập nhật thông tin người dùng "{user.username}".')
                        return redirect('user_list')
                        
                except Exception as e:
                    messages.error(request, f'Có lỗi xảy ra: {str(e)}')
        
        elif action == 'reset_password':
            new_password = request.POST.get('new_password', '')
            confirm_password = request.POST.get('confirm_password', '')
            
            if not new_password:
                messages.error(request, 'Vui lòng nhập mật khẩu mới.')
            elif new_password != confirm_password:
                messages.error(request, 'Mật khẩu xác nhận không khớp.')
            else:
                try:
                    user.set_password(new_password)
                    user.save()
                    messages.success(request, f'Đã reset mật khẩu cho người dùng "{user.username}".')
                    return redirect('user_list')
                except Exception as e:
                    messages.error(request, f'Có lỗi xảy ra: {str(e)}')
    
    # Lấy danh sách user groups
    user_groups = UserGroup.objects.all().order_by('user_group_name')
    
    # Xác định vai trò hiện tại để hiển thị nếu không tìm thấy LibraryUser
    current_role_display = 'Quản lý' if user.is_superuser else ('Thủ thư' if user.is_staff else 'Người dùng')
    if library_user and library_user.user_group:
        current_role_display = library_user.user_group.user_group_name
    
    context = {
        'edit_user': user,
        'user_groups': user_groups,
        'current_group_id': current_group_id,
        'current_role': current_role_display,
        'page_title': f'Chỉnh sửa người dùng: {user.username}'
    }
    
    return render(request, 'app/users/user_form.html', context)


@permission_required('Quản lý người dùng', 'delete')
def user_delete_view(request, user_id):
    """
    Xóa người dùng - Chỉ dành cho Quản lý
    
    Lưu ý:
    - Không thể xóa chính mình
    - Xóa vĩnh viễn (cẩn thận!)
    """
    user = get_object_or_404(User, id=user_id)
    
    # Không cho phép xóa chính mình
    if user.id == request.user.id:
        messages.error(request, 'Không thể xóa tài khoản của chính bạn!')
        return redirect('user_list')
    
    if request.method == 'POST':
        username = user.username
        try:
            user.delete()
            messages.success(request, f'Đã xóa người dùng "{username}" khỏi hệ thống.')
        except Exception as e:
            messages.error(request, f'Có lỗi xảy ra khi xóa: {str(e)}')
        
        return redirect('user_list')
    
    context = {
        'delete_user': user,
        'page_title': 'Xác nhận xóa người dùng'
    }
    
    return render(request, 'app/users/user_delete_confirm.html', context)


# ==================== ĐĂNG KÝ TÀI KHOẢN ====================

def register_view(request):
    """
    Đăng ký tài khoản mới
    
    - Tài khoản mới mặc định: is_active=False (cần Quản lý kích hoạt)
    - Vai trò mặc định: Thủ thư (is_staff=True, is_superuser=False)
    - Sau khi đăng ký, hiển thị thông báo chờ duyệt
    """
    # Nếu đã đăng nhập, chuyển về trang chủ
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        
        # Validation
        if not username or not email or not password:
            messages.error(request, 'Vui lòng điền đầy đủ thông tin bắt buộc.')
        elif password != password_confirm:
            messages.error(request, 'Mật khẩu xác nhận không khớp.')
        elif len(password) < 8:
            messages.error(request, 'Mật khẩu phải có ít nhất 8 ký tự.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, f'Tên đăng nhập "{username}" đã tồn tại.')
        elif User.objects.filter(email=email).exists():
            messages.error(request, f'Email "{email}" đã được sử dụng.')
        else:
            try:
                with transaction.atomic():
                    # Tạo user mới (CHƯA KÍCH HOẠT)
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password,
                        first_name=first_name,
                        last_name=last_name,
                        is_active=False,  # Cần Quản lý kích hoạt
                        is_staff=True,    # Mặc định là Thủ thư
                        is_superuser=False
                    )
                    
                    messages.success(
                        request,
                        f'Đăng ký tài khoản "{username}" thành công! '
                        'Vui lòng chờ quản trị viên kích hoạt tài khoản.'
                    )
                    return redirect('login')
                    
            except Exception as e:
                messages.error(request, f'Có lỗi xảy ra: {str(e)}')
    
    context = {
        'page_title': 'Đăng ký tài khoản'
    }
    
    return render(request, 'accounts/register.html', context)


# ==================== BOOK DETAIL & EDIT ====================

@permission_required('Quản lý kho sách', 'view')
def book_detail_view(request, book_id):
    """
    Xem chi tiết sách
    Hiển thị thông tin đầy đủ về sách và cho phép chuyển sang chỉnh sửa
    """
    book = get_object_or_404(Book, id=book_id)
    
    # Lấy thông tin về tình trạng mượn
    book_items = BookItem.objects.filter(book=book)
    borrowed_items = book_items.filter(is_borrowed=True)
    
    # Check permissions for actions
    from .decorators import check_permission
    can_edit_books = check_permission(request.user, 'Quản lý kho sách', 'change')
    can_delete_books = check_permission(request.user, 'Quản lý kho sách', 'delete')
    
    context = {
        'book': book,
        'book_items': book_items,
        'borrowed_count': borrowed_items.count(),
        'available_count': book_items.filter(is_borrowed=False).count(),
        'can_edit_books': can_edit_books,
        'can_delete_books': can_delete_books,
        'page_title': f'Chi tiết sách - {book.book_title.book_title}'
    }
    
    return render(request, 'app/books/book_detail.html', context)


@permission_required('Quản lý kho sách', 'delete')
def book_delete_view(request, book_id):
    """
    View xóa sách
    
    Business Rules:
    - Chỉ xóa được nếu không có book item nào đang được mượn
    - Yêu cầu quyền DELETE trong chức năng "Quản lý kho sách"
    - Sẽ xóa cả BookTitle nếu đây là Book cuối cùng của BookTitle đó
    """
    book = get_object_or_404(Book, id=book_id)
    
    if request.method == 'POST':
        try:
            # Kiểm tra book items đang được mượn
            borrowed_items = BookItem.objects.filter(
                book=book,
                is_borrowed=True
            )
            borrowed_count = borrowed_items.count()
            
            if borrowed_count > 0:
                messages.error(
                    request,
                    f'Không thể xóa sách "{book.book_title.book_title}" vì còn {borrowed_count} cuốn đang được mượn. '
                    'Vui lòng đợi tất cả các cuốn sách được trả về trước khi xóa.'
                )
                return redirect('book_detail', book_id=book.id)
            
            # Thực hiện xóa
            book_title_name = book.book_title.book_title
            book_title = book.book_title
            
            # Xóa book (cascade sẽ xóa book items)
            book.delete()
            
            # Kiểm tra xem BookTitle còn Book nào không
            # Nếu không còn Book nào, có thể cân nhắc xóa BookTitle
            # (Tùy business logic - hiện tại giữ lại BookTitle)
            
            messages.success(
                request,
                f'Đã xóa sách "{book_title_name}" thành công!'
            )
            return redirect('book_search')
            
        except Exception as e:
            messages.error(request, f'Có lỗi xảy ra khi xóa sách: {str(e)}')
            return redirect('book_detail', book_id=book.id)
    
    # GET: Hiển thị trang xác nhận xóa
    # Lấy thông tin về tình trạng mượn
    book_items = BookItem.objects.filter(book=book)
    borrowed_items = book_items.filter(is_borrowed=True)
    
    # Đếm số phiếu nhập liên quan
    import_count = BookImportDetail.objects.filter(book=book).count()
    
    context = {
        'book': book,
        'total_items': book_items.count(),
        'borrowed_count': borrowed_items.count(),
        'available_count': book_items.filter(is_borrowed=False).count(),
        'import_count': import_count,
        'can_delete': borrowed_items.count() == 0,
        'page_title': f'Xóa sách - {book.book_title.book_title}'
    }
    
    return render(request, 'app/books/book_delete_confirm.html', context)


@permission_required('Quản lý kho sách', 'edit')
def book_edit_view(request, book_id):
    """
    Chỉnh sửa thông tin sách - Phiên bản đầy đủ
    Cho phép chỉnh sửa: tất cả các trường bao gồm thể loại, tác giả, đơn giá, năm XB, NXB
    """
    book = get_object_or_404(Book, id=book_id)
    is_admin = request.user.is_superuser
    
    if request.method == 'POST':
        form = BookEditForm(request.POST, instance=book, is_admin=is_admin)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # 1. Lưu các trường thuộc Book (từ form)
                    form.save()
                    
                    # 2. Lưu các trường Book bổ sung (từ POST data)
                    book.unit_price = int(request.POST.get('unit_price', book.unit_price))
                    book.publish_year = int(request.POST.get('publish_year', book.publish_year))
                    book.publisher = request.POST.get('publisher', book.publisher).strip()
                    book.save()
                    
                    # 3. Cập nhật BookTitle
                    book_title = book.book_title
                    
                    # 3a. Cập nhật mô tả
                    description = request.POST.get('description', '').strip()
                    book_title.description = description
                    
                    # 3b. Cập nhật thể loại
                    category_id = request.POST.get('category')
                    if category_id:
                        new_category = Category.objects.get(id=int(category_id))
                        book_title.category = new_category
                    
                    book_title.save()
                    
                    # 3c. Cập nhật tác giả
                    # Đọc existing authors từ form
                    author_ids = request.POST.getlist('authors', [])
                    existing_author_ids = [aid for aid in author_ids if aid and not aid.startswith('new_')]
                    
                    # Đọc new authors từ hidden input
                    new_author_names_str = request.POST.get('new_author_names', '')
                    new_author_names = [name.strip() for name in new_author_names_str.split('|||') if name.strip()]
                    
                    # Xóa tất cả author hiện tại và thêm lại
                    AuthorDetail.objects.filter(book_title=book_title).delete()
                    
                    # Thêm existing authors
                    for aid in existing_author_ids:
                        try:
                            author = Author.objects.get(id=int(aid))
                            AuthorDetail.objects.create(author=author, book_title=book_title)
                        except (Author.DoesNotExist, ValueError):
                            pass
                    
                    # Tạo và thêm new authors
                    for name in new_author_names:
                        # Case-insensitive lookup để tránh trùng lặp
                        existing = Author.objects.filter(author_name__iexact=name).first()
                        if existing:
                            author = existing
                        else:
                            author = Author.objects.create(author_name=name)
                        AuthorDetail.objects.get_or_create(author=author, book_title=book_title)
                    
                    messages.success(request, f'Cập nhật sách "{book.book_title.book_title}" thành công!')
                    return redirect('book_detail', book_id=book.id)
                    
            except Exception as e:
                messages.error(request, f'Có lỗi xảy ra: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    else:
        form = BookEditForm(instance=book, is_admin=is_admin)
    
    # Lấy danh sách NXB có sẵn
    existing_publishers = list(Book.objects.values_list('publisher', flat=True).distinct().order_by('publisher'))
    
    # Kiểm tra quyền xoá sách
    from .decorators import check_permission
    can_delete_book = request.user.is_superuser or check_permission(request.user, 'Quản lý kho sách', 'delete')
    
    context = {
        'form': form,
        'book': book,
        'is_admin': is_admin,
        'can_delete_book': can_delete_book,
        'categories': Category.objects.all().order_by('category_name'),
        'all_authors': Author.objects.all().order_by('author_name'),
        'existing_publishers': existing_publishers,
        'page_title': f'Chỉnh sửa sách - {book.book_title.book_title}'
    }
    
    return render(request, 'app/books/book_edit.html', context)


@permission_required('Quản lý kho sách', 'delete')
def book_delete_view(request, book_id):
    """
    Xoá sách khỏi hệ thống
    
    Điều kiện xoá:
    - Sách không đang được mượn (không có BookItem nào is_borrowed=True)
    """
    book = get_object_or_404(Book, id=book_id)
    book_title_name = book.book_title.book_title
    
    # Kiểm tra điều kiện xoá
    borrowed_items = BookItem.objects.filter(book=book, is_borrowed=True)
    if borrowed_items.exists():
        borrowed_count = borrowed_items.count()
        messages.error(
            request, 
            f'Không thể xoá sách "{book_title_name}" vì đang có {borrowed_count} cuốn được mượn. '
            'Vui lòng chờ độc giả trả sách trước khi xoá.'
        )
        return redirect('book_detail', book_id=book.id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Xoá các BookItem liên quan
                BookItem.objects.filter(book=book).delete()
                
                # Xoá các BookImportDetail liên quan
                BookImportDetail.objects.filter(book=book).delete()
                
                # Xoá sách
                book.delete()
                
                messages.success(request, f'Đã xoá sách "{book_title_name}" thành công!')
                return redirect('book_search')
                
        except Exception as e:
            messages.error(request, f'Có lỗi xảy ra khi xoá sách: {str(e)}')
            return redirect('book_detail', book_id=book_id)
    
    # GET request - hiển thị trang xác nhận
    context = {
        'book': book,
        'book_items_count': BookItem.objects.filter(book=book).count(),
        'page_title': f'Xoá sách - {book_title_name}'
    }
    
    return render(request, 'app/books/book_delete.html', context)

@manager_required
def reader_type_list_view(request):
    """
    Danh sách loại độc giả
    """
    reader_types = ReaderType.objects.annotate(
        reader_count=Count('readers')
    ).order_by('reader_type_name')
    
    context = {
        'reader_types': reader_types,
        'page_title': 'Quản lý loại độc giả'
    }
    
    return render(request, 'app/settings/reader_type_list.html', context)


@manager_required
def reader_type_create_view(request):
    """
    Thêm loại độc giả mới
    """
    if request.method == 'POST':
        form = ReaderTypeForm(request.POST)
        if form.is_valid():
            reader_type = form.save()
            messages.success(request, f'Đã thêm loại độc giả "{reader_type.reader_type_name}" thành công!')
            return redirect('reader_type_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    else:
        form = ReaderTypeForm()
    
    context = {
        'form': form,
        'page_title': 'Thêm loại độc giả mới'
    }
    
    return render(request, 'app/settings/reader_type_form.html', context)


@manager_required
def reader_type_edit_view(request, reader_type_id):
    """
    Chỉnh sửa loại độc giả
    """
    reader_type = get_object_or_404(ReaderType, id=reader_type_id)
    
    if request.method == 'POST':
        form = ReaderTypeForm(request.POST, instance=reader_type)
        if form.is_valid():
            form.save()
            messages.success(request, f'Đã cập nhật loại độc giả "{reader_type.reader_type_name}" thành công!')
            return redirect('reader_type_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    else:
        form = ReaderTypeForm(instance=reader_type)
    
    context = {
        'form': form,
        'reader_type': reader_type,
        'page_title': f'Chỉnh sửa loại độc giả - {reader_type.reader_type_name}'
    }
    
    return render(request, 'app/settings/reader_type_form.html', context)


@manager_required
def reader_type_delete_view(request, reader_type_id):
    """
    Xóa loại độc giả
    Không cho phép xóa nếu còn độc giả thuộc loại này
    """
    reader_type = get_object_or_404(ReaderType, id=reader_type_id)
    
    # Kiểm tra có độc giả nào thuộc loại này không
    reader_count = Reader.objects.filter(reader_type=reader_type).count()
    
    if request.method == 'POST':
        if reader_count > 0:
            messages.error(
                request, 
                f'Không thể xóa loại độc giả "{reader_type.reader_type_name}" vì còn {reader_count} độc giả thuộc loại này.'
            )
        else:
            name = reader_type.reader_type_name
            reader_type.delete()
            messages.success(request, f'Đã xóa loại độc giả "{name}" thành công!')
        return redirect('reader_type_list')
    
    context = {
        'reader_type': reader_type,
        'reader_count': reader_count,
        'page_title': f'Xóa loại độc giả - {reader_type.reader_type_name}'
    }
    
    return render(request, 'app/settings/reader_type_delete.html', context)


# ==================== PERMISSION MANAGEMENT VIEWS ====================

@permission_required('Quản lý quyền', 'view')
def user_group_list_view(request):
    """Danh sách nhóm người dùng"""
    user_groups = UserGroup.objects.all().order_by('user_group_name')
    
    context = {
        'user_groups': user_groups,
        'page_title': 'Quản lý nhóm người dùng'
    }
    return render(request, 'app/permissions/user_group_list.html', context)


@permission_required('Quản lý quyền', 'add')
def user_group_create_view(request):
    """Tạo nhóm người dùng mới"""
    if request.method == 'POST':
        form = UserGroupForm(request.POST)
        if form.is_valid():
            user_group = form.save()
            messages.success(request, f'Đã tạo nhóm "{user_group.user_group_name}" thành công!')
            return redirect('user_group_list')
    else:
        form = UserGroupForm()
    
    context = {
        'form': form,
        'page_title': 'Thêm nhóm người dùng',
        'is_edit': False
    }
    return render(request, 'app/permissions/user_group_form.html', context)


@permission_required('Quản lý quyền', 'edit')
def user_group_edit_view(request, group_id):
    """Sửa nhóm người dùng"""
    user_group = get_object_or_404(UserGroup, id=group_id)
    
    if request.method == 'POST':
        form = UserGroupForm(request.POST, instance=user_group)
        if form.is_valid():
            form.save()
            messages.success(request, f'Đã cập nhật nhóm "{user_group.user_group_name}" thành công!')
            return redirect('user_group_list')
    else:
        form = UserGroupForm(instance=user_group)
    
    context = {
        'form': form,
        'user_group': user_group,
        'page_title': f'Sửa nhóm - {user_group.user_group_name}',
        'is_edit': True
    }
    return render(request, 'app/permissions/user_group_form.html', context)


@permission_required('Quản lý quyền', 'delete')
def user_group_delete_view(request, group_id):
    """Xóa nhóm người dùng"""
    user_group = get_object_or_404(UserGroup, id=group_id)
    
    # Đếm số user thuộc nhóm này
    from .models import LibraryUser
    user_count = LibraryUser.objects.filter(user_group=user_group).count()
    
    if request.method == 'POST':
        if user_count > 0:
            messages.error(request, f'Không thể xóa nhóm "{user_group.user_group_name}" vì còn {user_count} người dùng!')
        else:
            name = user_group.user_group_name
            user_group.delete()
            messages.success(request, f'Đã xóa nhóm "{name}" thành công!')
        return redirect('user_group_list')
    
    context = {
        'user_group': user_group,
        'user_count': user_count,
        'page_title': f'Xóa nhóm - {user_group.user_group_name}'
    }
    return render(request, 'app/permissions/user_group_delete.html', context)


@permission_required('Quản lý quyền', 'view')
def function_list_view(request):
    """Danh sách chức năng"""
    functions = Function.objects.all().order_by('function_name')
    
    context = {
        'functions': functions,
        'page_title': 'Quản lý chức năng hệ thống'
    }
    return render(request, 'app/permissions/function_list.html', context)


@permission_required('Quản lý quyền', 'add')
def function_create_view(request):
    """Tạo chức năng mới"""
    if request.method == 'POST':
        form = FunctionForm(request.POST)
        if form.is_valid():
            function = form.save()
            messages.success(request, f'Đã tạo chức năng "{function.function_name}" thành công!')
            return redirect('function_list')
    else:
        form = FunctionForm()
    
    context = {
        'form': form,
        'page_title': 'Thêm chức năng',
        'is_edit': False
    }
    return render(request, 'app/permissions/function_form.html', context)


@permission_required('Quản lý quyền', 'edit')
def function_edit_view(request, function_id):
    """Sửa chức năng"""
    function = get_object_or_404(Function, id=function_id)
    
    if request.method == 'POST':
        form = FunctionForm(request.POST, instance=function)
        if form.is_valid():
            form.save()
            messages.success(request, f'Đã cập nhật chức năng "{function.function_name}" thành công!')
            return redirect('function_list')
    else:
        form = FunctionForm(instance=function)
    
    context = {
        'form': form,
        'function': function,
        'page_title': f'Sửa chức năng - {function.function_name}',
        'is_edit': True
    }
    return render(request, 'app/permissions/function_form.html', context)


@permission_required('Quản lý quyền', 'delete')
def function_delete_view(request, function_id):
    """Xóa chức năng"""
    function = get_object_or_404(Function, id=function_id)
    
    # Đếm số permission liên quan
    permission_count = Permission.objects.filter(function=function).count()
    
    if request.method == 'POST':
        name = function.function_name
        function.delete()
        messages.success(request, f'Đã xóa chức năng "{name}" thành công!')
        return redirect('function_list')
    
    context = {
        'function': function,
        'permission_count': permission_count,
        'page_title': f'Xóa chức năng - {function.function_name}'
    }
    return render(request, 'app/permissions/function_delete.html', context)


@permission_required('Quản lý quyền', 'edit')
def permission_matrix_view(request, group_id):
    """Quản lý quyền cho nhóm người dùng (ma trận quyền)"""
    user_group = get_object_or_404(UserGroup, id=group_id)
    functions = Function.objects.all().order_by('function_name')
    
    if request.method == 'POST':
        # Xử lý cập nhật quyền
        for function in functions:
            can_view = request.POST.get(f'view_{function.id}') == 'on'
            can_add = request.POST.get(f'add_{function.id}') == 'on'
            can_edit = request.POST.get(f'edit_{function.id}') == 'on'
            can_delete = request.POST.get(f'delete_{function.id}') == 'on'
            
            # Cập nhật hoặc tạo mới permission
            permission, created = Permission.objects.update_or_create(
                user_group=user_group,
                function=function,
                defaults={
                    'can_view': can_view,
                    'can_add': can_add,
                    'can_edit': can_edit,
                    'can_delete': can_delete
                }
            )
        
        messages.success(request, f'Đã cập nhật quyền cho nhóm "{user_group.user_group_name}" thành công!')
        return redirect('permission_matrix', group_id=group_id)
    
    # Lấy quyền hiện tại
    permissions = {}
    for p in Permission.objects.filter(user_group=user_group):
        permissions[str(p.function_id)] = {
            'can_view': p.can_view,
            'can_add': p.can_add,
            'can_edit': p.can_edit,
            'can_delete': p.can_delete
        }
    
    import json
    context = {
        'user_group': user_group,
        'functions': functions,
        'permissions': json.dumps(permissions),
        'page_title': f'Phân quyền - {user_group.user_group_name}'
    }
    return render(request, 'app/permissions/permission_matrix.html', context)

